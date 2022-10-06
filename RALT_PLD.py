#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
This experiment was created using PsychoPy3 Experiment Builder (v2020.2.10),
    on Mon Feb  7 10:28:54 2022
If you publish work using this script the most relevant publication is:

    Peirce J, Gray JR, Simpson S, MacAskill M, Höchenberger R, Sogo H, Kastman E, Lindeløv JK. (2019) 
        PsychoPy2: Experiments in behavior made easy Behav Res 51: 195. 
        https://doi.org/10.3758/s13428-018-01193-y

"""

from __future__ import absolute_import, division

from psychopy import locale_setup
from psychopy import prefs
from psychopy import sound, gui, visual, core, data, event, logging, clock
from psychopy.constants import (NOT_STARTED, STARTED, PLAYING, PAUSED,
                                STOPPED, FINISHED, PRESSED, RELEASED, FOREVER)

import numpy as np  # whole numpy lib is available, prepend 'np.'
from numpy import (sin, cos, tan, log, log10, pi, average,
                   sqrt, std, deg2rad, rad2deg, linspace, asarray)
from numpy.random import random, randint, normal, shuffle
import os  # handy system and path functions
import sys  # to get file system encoding

from psychopy.hardware import keyboard



# Ensure that relative paths start from the same directory as this script
_thisDir = os.path.dirname(os.path.abspath(__file__))
os.chdir(_thisDir)

# Store info about the experiment session
psychopyVersion = '2020.2.10'
expName = 'EmoSexCounterbalanced_fontfix'  # from the Builder filename that created this script
expInfo = {'participant': '', 'session': '001'}
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName
expInfo['psychopyVersion'] = psychopyVersion

# Data file name stem = absolute path + name; later add .psyexp, .csv, .log, etc
filename = _thisDir + os.sep + u'data/%s_%s_%s' %(expInfo['participant'], expName, expInfo['date'])

# An ExperimentHandler isn't essential but helps with data saving
thisExp = data.ExperimentHandler(name=expName, version='',
    extraInfo=expInfo, runtimeInfo=None,
    originPath='/Users/raimundbuehler/Desktop/RALT_PLD_english/RALT_PLD.py',
    savePickle=True, saveWideText=True,
    dataFileName=filename)
logging.console.setLevel(logging.WARNING)  # this outputs to the screen, not a file

endExpNow = False  # flag for 'escape' or other condition => quit the exp
frameTolerance = 0.001  # how close to onset before 'same' frame

# Start Code - component code to be run after the window creation

# Setup the Window
win = visual.Window(
    size=[1920, 1080], fullscr=False, screen=0, 
    winType='pyglet', allowGUI=True, allowStencil=False,
    monitor='OfficeMonitor', color=[-1,-1,-1], colorSpace='rgb',
    blendMode='avg', useFBO=True, 
    units='height')
# store frame rate of monitor if we can measure it
expInfo['frameRate'] = win.getActualFrameRate()
if expInfo['frameRate'] != None:
    frameDur = 1.0 / round(expInfo['frameRate'])
else:
    frameDur = 1.0 / 60.0  # could not measure, so guess

# create a default keyboard (e.g. to check for escape)
defaultKeyboard = keyboard.Keyboard()

# Initialize components for Routine "Welcome"
WelcomeClock = core.Clock()
import openpyxl
import numpy
import random
from numpy.random import choice

win.setColor('white')
font_choice = 'Menlo'

#Shuffle correct category(left/right) and write to excel file

cwd = os.getcwd()

#Randomize block conditions xl file
'''
xlfile_cond = "condBlocks.xlsx"
path_cond = os.path.join(cwd, xlfile_cond)

book_cond = openpyxl.load_workbook(filename = path_cond)

sheet = book_cond.active

list_to_shuffle_cond = ['social', 'nonsocial']
random.shuffle(list_to_shuffle_cond)
list_to_shuffle_cond = list_to_shuffle_cond + list_to_shuffle_cond

for index, element in enumerate(list_to_shuffle_cond, start=1):
    sheet.cell(row = index + 1, column = 1).value = element

book_cond.save(filename = "condBlocks.xlsx")
'''
#Funktion zur Zentrierung des Textes
def centerfy(str1):
    str2 = str1.split('\n')
    str3 = ''
    for line in str2:
        str3 = str3 + '\n' + line.center(30)
    return str3



#Zentrierter Text;
Welcome_heading = centerfy("Vielen Dank, dass Sie an diesem Experiment teilnehmen!")

Welcome = centerfy("Sie werden im Folgenden gebeten,\n verschiedene Zahlen zuzuordnen.\n"
'\n'
"Sie können dies tun, indem Sie die Pfeiltasten <- oder -> drücken.\n"
'\n'
"Ob Sie richtig gewählt haben\n wird Ihnen durch Bilder angezeigt.\n"
'\n'
"Wir beginnen mit einigen Probedurchgängen, \n damit Sie das Experiment besser verstehen.\n\n"
'Drücken Sie eine beliebige Taste, um fortzufahren!')

instruction = centerfy("Sie werden nun gleich eine Zahl \n im unteren Bereich des Bildschirms sehen."
'\n\n'
'Ihre Aufgabe ist es, die Zahl mittels der Pfeiltasten \n entweder A oder B zuzuordnen.'
'\n\n'
'Die richtige Zuordnung ist völlig zufällig.\n'
'Sie müssen also zunächst einfach raten.'
'\n\n'
'Zusätzlich werden Sie über der Zahl ein Bild sehen,\n'
'in diesem Fall ein neutrales Gesicht.'
'\n\n'
'Ob Sie richtig gewählt haben, \n wird Ihnen durch eine Veränderung des Bildes angezeigt.'
'\n\n\n'
'Versuchen wir es mit einem ersten Durchgang!')

if choice([True, False]):
    right_cat = "A"
    left_cat = "B"
else:
    left_cat = "A"
    right_cat = "B"
Welcome_head = visual.TextStim(win=win, name='Welcome_head',
    text='Thank you for participating in this experiment!',
    font=font_choice,
    pos=(0, 0.3), height=0.025, wrapWidth=70, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-2.0);
Welcome_text = visual.TextStim(win=win, name='Welcome_text',
    text="You will be asked to assign different numbers to either A or B.\n\nYou can do this by pressing the <- or -> arrow keys.\nWhether you have chosen correctly will be indicated by pictures.\n\nWe'll start with some trial runs,\nso you can better understand the experiment.\n\n\nPress any key to continue!",
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=70, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-3.0);
key_resp_2 = keyboard.Keyboard()

# Initialize components for Routine "instr_train_s"
instr_train_sClock = core.Clock()
text_3 = visual.TextStim(win=win, name='text_3',
    text="You are about to see a number at the bottom of the screen. \nYour task is to assign the number to either A or B using the arrow keys.\n\nThe correct assignment is completely random\nso you just have to guess at first.\n\nAdditionally, you will see a picture above the number.\nWhether you have chosen correctly will be indicated by a change in the picture.\n\n\nLet's try a first round!",
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=70, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=0.0);
key_resp_3 = keyboard.Keyboard()

# Initialize components for Routine "trial"
trialClock = core.Clock()
fix_cross = visual.TextStim(win=win, name='fix_cross',
    text='+',
    font=font_choice,
    pos=(0, 0), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
neutral_still = visual.ImageStim(
    win=win,
    name='neutral_still', units='height', 
    image='sin', mask=None,
    ori=0, pos=[0, 0.225], size=(0.5, 0.5),
    color=[1,1,1], colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=512, interpolate=True, depth=-2.0)
response_training = keyboard.Keyboard()
Stimulus = visual.TextStim(win=win, name='Stimulus',
    text='default text',
    font=font_choice,
    pos=(0, -0.15), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
left_disp = visual.TextStim(win=win, name='left_disp',
    text='default text',
    font=font_choice,
    pos=(-0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='grey', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-5.0);
right_disp = visual.TextStim(win=win, name='right_disp',
    text='default text',
    font=font_choice,
    pos=(0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='grey', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-6.0);

# Initialize components for Routine "feedback_train"
feedback_trainClock = core.Clock()
feedback_miss = visual.TextStim(win=win, name='feedback_miss',
    text='PLEASE RESPOND FASTER!',
    font=font_choice,
    pos=(0, 0.225), height=0.05, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=0.0);
left_disp_fb_2 = visual.TextStim(win=win, name='left_disp_fb_2',
    text='default text',
    font=font_choice,
    pos=(-0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
right_disp_fb_2 = visual.TextStim(win=win, name='right_disp_fb_2',
    text='default text',
    font=font_choice,
    pos=(0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-5.0);
Stimulus_fb_2 = visual.TextStim(win=win, name='Stimulus_fb_2',
    text='default text',
    font=font_choice,
    pos=(0, -0.15), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-6.0);

# Initialize components for Routine "Alles_klar"
Alles_klarClock = core.Clock()
Allesklar = centerfy(('Wie Sie gesehen haben\n verändert sich das Bild entsprechend Ihrer Antwort.\n\n'
'Fröhliche Gesichter zeigen eine richtige Antwort an,\n ärgerliche Gesichter eine Falsche.\n\n'
'Später sollen Sie versuchen,\n sich die richtige Zuordnung der Zahl genau einzuprägen. \n\n'
'Bild und Zahl sind einander zufällig zugeordnet\n und werden später neu kombiniert.\n'
'Sie müssen sich beim Einprägen also auf die Zahl konzentrieren!\n\n'
'Drücken Sie eine beliebige Taste, um fortzufahren!'))
text_4 = visual.TextStim(win=win, name='text_4',
    text='As you have seen, the picture changes according to your answer.\nHappy faces indicate a correct answer, \nwhile angry faces indicate a wrong answer.\n\nLater on, you should try to remember the correct assignment of the number. \n\nPress any key to continue!',
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=70, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
key_resp_4 = keyboard.Keyboard()

# Initialize components for Routine "instr_train_ns"
instr_train_nsClock = core.Clock()
Allesklar = centerfy(('Im nächsten Durchgang werden Sie etwas andere Bilder sehen.\n\n'
'Das Prinzip bleibt aber gleich:\n'
' Das Bild verändert sich entsprechend einer richtigen/falschen Antwort.\n'
'Probieren wir es aus!'))
text = visual.TextStim(win=win, name='text',
    text="In the next pass, you will see different pictures. \n\nBut the principle remains the same\nThe picture changes according to a correct/wrong answer.\n\nLet's try it out!",
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=70, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
key_resp_5 = keyboard.Keyboard()

# Initialize components for Routine "trial"
trialClock = core.Clock()
fix_cross = visual.TextStim(win=win, name='fix_cross',
    text='+',
    font=font_choice,
    pos=(0, 0), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
neutral_still = visual.ImageStim(
    win=win,
    name='neutral_still', units='height', 
    image='sin', mask=None,
    ori=0, pos=[0, 0.225], size=(0.5, 0.5),
    color=[1,1,1], colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=512, interpolate=True, depth=-2.0)
response_training = keyboard.Keyboard()
Stimulus = visual.TextStim(win=win, name='Stimulus',
    text='default text',
    font=font_choice,
    pos=(0, -0.15), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
left_disp = visual.TextStim(win=win, name='left_disp',
    text='default text',
    font=font_choice,
    pos=(-0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='grey', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-5.0);
right_disp = visual.TextStim(win=win, name='right_disp',
    text='default text',
    font=font_choice,
    pos=(0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='grey', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-6.0);

# Initialize components for Routine "feedback_train"
feedback_trainClock = core.Clock()
feedback_miss = visual.TextStim(win=win, name='feedback_miss',
    text='PLEASE RESPOND FASTER!',
    font=font_choice,
    pos=(0, 0.225), height=0.05, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=0.0);
left_disp_fb_2 = visual.TextStim(win=win, name='left_disp_fb_2',
    text='default text',
    font=font_choice,
    pos=(-0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
right_disp_fb_2 = visual.TextStim(win=win, name='right_disp_fb_2',
    text='default text',
    font=font_choice,
    pos=(0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-5.0);
Stimulus_fb_2 = visual.TextStim(win=win, name='Stimulus_fb_2',
    text='default text',
    font=font_choice,
    pos=(0, -0.15), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-6.0);

# Initialize components for Routine "Alles_Klar_2"
Alles_Klar_2Clock = core.Clock()
Allesklar = centerfy('Wie Sie gesehen haben,\n ändert sich das Bild wieder entsprechend Ihrer Antwort.\n\n'
'Bunte, bewegte Bilder zeigen eine richtige Antwort an,\n bei einer falschen Antwort wird das Bild grau und verschwommen.\n\n'
'Ganz so einfach wird es aber im richtigen Experiment leider nicht.\n'
'Sie werden sich nicht immer auf die Bilder verlassen können.\n'
'Stattdessen bekommen Sie bei einer richtigen Antwort MEISTENS die fröhlichen Gesichter bzw. bunten Fraktale zu sehen. \n'
'Das heißt, selbst wenn Sie richtig geantwortet haben, \n' 
'kann es manchmal sein, dass Sie trotzdem ein ärgerliches Gesicht bzw. ein graues Fraktal sehen \n'
'Im nächsten Durchgang wird durch Text angezeigt, ob Sie richtig oder falsch geantwortet haben. \n'
'Achten Sie darauf, dass Text und Bild meistens, aber nicht immer übereinstimmen.')
Allesklartext_2 = visual.TextStim(win=win, name='Allesklartext_2',
    text="As you have seen, the image changes again according to your answer.\nThis time, tickmarks indicate a correct answer, while a cross indicates a wrong answer. \n\nBut it won't be quite that simple in the real experiment!\nYou will not always be able to rely on the pictures.\n\nInstead, if you answer correctly, you will\nMOSTLY see the happy faces or tickmarks.\n\nThat is, even if you answered correctly,\nsometimes you may still see a frowning face or a cross.\n\nIn the next pass, text will indicate whether you answered correctly or incorrectly.\n\nNote that text and image mostly, but not always, match.",
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=70, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
Allesklar_resp_2 = keyboard.Keyboard()

# Initialize components for Routine "trial"
trialClock = core.Clock()
fix_cross = visual.TextStim(win=win, name='fix_cross',
    text='+',
    font=font_choice,
    pos=(0, 0), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
neutral_still = visual.ImageStim(
    win=win,
    name='neutral_still', units='height', 
    image='sin', mask=None,
    ori=0, pos=[0, 0.225], size=(0.5, 0.5),
    color=[1,1,1], colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=512, interpolate=True, depth=-2.0)
response_training = keyboard.Keyboard()
Stimulus = visual.TextStim(win=win, name='Stimulus',
    text='default text',
    font=font_choice,
    pos=(0, -0.15), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
left_disp = visual.TextStim(win=win, name='left_disp',
    text='default text',
    font=font_choice,
    pos=(-0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='grey', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-5.0);
right_disp = visual.TextStim(win=win, name='right_disp',
    text='default text',
    font=font_choice,
    pos=(0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='grey', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-6.0);

# Initialize components for Routine "feedback_prob"
feedback_probClock = core.Clock()
#Initialize color
feedback = ''
left_color = ''
right_color = ''
feedback_miss2_3 = visual.TextStim(win=win, name='feedback_miss2_3',
    text='PLEASE RESPOND FASTER!',
    font=font_choice,
    pos=(0, 0.225), height=0.05, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
left_disp_fb_4 = visual.TextStim(win=win, name='left_disp_fb_4',
    text='default text',
    font=font_choice,
    pos=(-0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
right_disp_fb_4 = visual.TextStim(win=win, name='right_disp_fb_4',
    text='default text',
    font=font_choice,
    pos=(0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-5.0);
Stimulus_fb_4 = visual.TextStim(win=win, name='Stimulus_fb_4',
    text='default text',
    font=font_choice,
    pos=(0, -0.15), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-6.0);
Prob_feedback = visual.TextStim(win=win, name='Prob_feedback',
    text='default text',
    font=font_choice,
    pos=(0, -0.25), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-7.0);

# Initialize components for Routine "Alles_Klar_3"
Alles_Klar_3Clock = core.Clock()
Allesklar = centerfy('Beim letzten Durchgang haben Bild und Text nicht übereingestimmt!\n Dies wird im richtigen Experiment aber nur selten passieren.\n\n'
'Den Text unten werden Sie im Folgenden nicht mehr angezeigt bekommen.\n\n'
'Versuchen Sie einfach, so gut es geht, die richtige Antwort zu geben!\n')
Allesklartext = visual.TextStim(win=win, name='Allesklartext',
    text="Did you notice? In the last pass, the picture and text didn't match!\n\nBut this will rarely happen in the real experiment.\n\nYou won't see the text below in the following.\n\nJust try to give the right answer as best you can!",
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=70, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
Allesklar_resp = keyboard.Keyboard()

# Initialize components for Routine "BlockCode"
BlockCodeClock = core.Clock()
#Set up order of Blocks (beginning randomized, then alternating)
Blocklist = ["social", "nonsocial"]
random.shuffle(Blocklist)
Blocklist = Blocklist*2

#Set up 16 2-digit stimuli, no replacement
fN = random.sample(range(10, 100), 16)
numbarray = [fN[0:4],
             fN[4:8],
             fN[8:12],
             fN[12:16]]

#Set up list with correct category (4 per cycle * 4 blocks, randomized)
CorrCat = []

for x in range(4):
    ABlist = ["A", "A", "B", "B"]
    random.shuffle(ABlist)
    CorrCat.append(ABlist)

#flatten list
CorrCat_new = [item for sublist in CorrCat for item in sublist]
#Set up dict with correct category for lookup
CorrDict = dict(zip(fN, CorrCat_new))
print("CorrDict:", CorrDict)

# Initialize components for Routine "LateralizationByCycle"
LateralizationByCycleClock = core.Clock()
CycleText = ''
CycleText1 = visual.TextStim(win=win, name='CycleText1',
    text='We are now starting the first block.\n\nIn the first pass you must\nguess the assignment with the arrow keys.\nTry to memorize the correct assignment!\n\nIn total, you will see 4 different numbers.\nAfter that, there will be a short pause.\n\n\nPress any key,\nto start the first block!',
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
CycleText2 = visual.TextStim(win=win, name='CycleText2',
    text='PAUSE\n\nPress any key,\nto start the next pass.\n\n(A and B might have switched places!)',
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-2.0);
CycleText3 = visual.TextStim(win=win, name='CycleText3',
    text='default text',
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-3.0);
CycleText4 = visual.TextStim(win=win, name='CycleText4',
    text='You have now completed a full block.\n\nIn the next block you will see different images.\n\nInstead of checkmarks/crosses you will now see faces (or vice versa).\n\nThe principle remains the same.\n\nThere are 4 blocks in total.\n\nWhen you are ready\npress any button,\nto start the next block.',
    font=font_choice,
    pos=(0, 0), height=0.025, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
key_resp_6 = keyboard.Keyboard()

# Initialize components for Routine "trial_2"
trial_2Clock = core.Clock()
fix_cross_2 = visual.TextStim(win=win, name='fix_cross_2',
    text='+',
    font=font_choice,
    pos=(0, 0), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
new_neutral_2_jpg = visual.ImageStim(
    win=win,
    name='new_neutral_2_jpg', units='height', 
    image='sin', mask=None,
    ori=0, pos=[0, 0.225], size=(0.5, 0.5),
    color=[1,1,1], colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=512, interpolate=True, depth=-2.0)
response_training_2 = keyboard.Keyboard()
Stimulus_2 = visual.TextStim(win=win, name='Stimulus_2',
    text='default text',
    font=font_choice,
    pos=(0, -0.15), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
left_disp_2 = visual.TextStim(win=win, name='left_disp_2',
    text='default text',
    font=font_choice,
    pos=(-0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='grey', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-5.0);
right_disp_2 = visual.TextStim(win=win, name='right_disp_2',
    text='default text',
    font=font_choice,
    pos=(0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='grey', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-6.0);

# Initialize components for Routine "feedback"
feedbackClock = core.Clock()
#Initialize color
feedback = ''
left_color = ''
right_color = ''
feedback_miss2 = visual.TextStim(win=win, name='feedback_miss2',
    text='PLEASE RESPOND FASTER!',
    font=font_choice,
    pos=(0, 0.225), height=0.05, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
left_disp_fb = visual.TextStim(win=win, name='left_disp_fb',
    text='default text',
    font=font_choice,
    pos=(-0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-4.0);
right_disp_fb = visual.TextStim(win=win, name='right_disp_fb',
    text='default text',
    font=font_choice,
    pos=(0.5, 0), height=0.1, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-5.0);
Stimulus_fb = visual.TextStim(win=win, name='Stimulus_fb',
    text='default text',
    font=font_choice,
    pos=(0, -0.15), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-6.0);

# Initialize components for Routine "Intertrial_Interval"
Intertrial_IntervalClock = core.Clock()
text_2 = visual.TextStim(win=win, name='text_2',
    text='+',
    font=font_choice,
    pos=(0, 0), height=0.1, wrapWidth=None, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=0.0);

# Initialize components for Routine "LatCounter"
LatCounterClock = core.Clock()

# Initialize components for Routine "BlockCounter"
BlockCounterClock = core.Clock()

# Initialize components for Routine "Thanks"
ThanksClock = core.Clock()
Thank = visual.TextStim(win=win, name='Thank',
    text='Thank you for your participation!',
    font=font_choice,
    pos=(0, 0), height=0.035, wrapWidth=70, ori=0, 
    color=color_choice, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=0.0);
end = keyboard.Keyboard()

# Create some handy timers
globalClock = core.Clock()  # to track the time since experiment started
routineTimer = core.CountdownTimer()  # to track time remaining of each (non-slip) routine 

# ------Prepare to start Routine "Welcome"-------
continueRoutine = True
# update component parameters for each repeat
Welcome_head.bold = True

key_resp_2.keys = []
key_resp_2.rt = []
_key_resp_2_allKeys = []
# keep track of which components have finished
WelcomeComponents = [Welcome_head, Welcome_text, key_resp_2]
for thisComponent in WelcomeComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
WelcomeClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "Welcome"-------
while continueRoutine:
    # get current time
    t = WelcomeClock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=WelcomeClock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *Welcome_head* updates
    if Welcome_head.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        Welcome_head.frameNStart = frameN  # exact frame index
        Welcome_head.tStart = t  # local t and not account for scr refresh
        Welcome_head.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(Welcome_head, 'tStartRefresh')  # time at next scr refresh
        Welcome_head.setAutoDraw(True)
    
    # *Welcome_text* updates
    if Welcome_text.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        Welcome_text.frameNStart = frameN  # exact frame index
        Welcome_text.tStart = t  # local t and not account for scr refresh
        Welcome_text.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(Welcome_text, 'tStartRefresh')  # time at next scr refresh
        Welcome_text.setAutoDraw(True)
    
    # *key_resp_2* updates
    waitOnFlip = False
    if key_resp_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp_2.frameNStart = frameN  # exact frame index
        key_resp_2.tStart = t  # local t and not account for scr refresh
        key_resp_2.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp_2, 'tStartRefresh')  # time at next scr refresh
        key_resp_2.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp_2.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp_2.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp_2.status == STARTED and not waitOnFlip:
        theseKeys = key_resp_2.getKeys(keyList=None, waitRelease=False)
        _key_resp_2_allKeys.extend(theseKeys)
        if len(_key_resp_2_allKeys):
            key_resp_2.keys = _key_resp_2_allKeys[-1].name  # just the last key pressed
            key_resp_2.rt = _key_resp_2_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in WelcomeComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "Welcome"-------
for thisComponent in WelcomeComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('Welcome_head.started', Welcome_head.tStartRefresh)
thisExp.addData('Welcome_head.stopped', Welcome_head.tStopRefresh)
thisExp.addData('Welcome_text.started', Welcome_text.tStartRefresh)
thisExp.addData('Welcome_text.stopped', Welcome_text.tStopRefresh)
# the Routine "Welcome" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# ------Prepare to start Routine "instr_train_s"-------
continueRoutine = True
# update component parameters for each repeat
key_resp_3.keys = []
key_resp_3.rt = []
_key_resp_3_allKeys = []
# keep track of which components have finished
instr_train_sComponents = [text_3, key_resp_3]
for thisComponent in instr_train_sComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
instr_train_sClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "instr_train_s"-------
while continueRoutine:
    # get current time
    t = instr_train_sClock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=instr_train_sClock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *text_3* updates
    if text_3.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        text_3.frameNStart = frameN  # exact frame index
        text_3.tStart = t  # local t and not account for scr refresh
        text_3.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(text_3, 'tStartRefresh')  # time at next scr refresh
        text_3.setAutoDraw(True)
    
    # *key_resp_3* updates
    waitOnFlip = False
    if key_resp_3.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp_3.frameNStart = frameN  # exact frame index
        key_resp_3.tStart = t  # local t and not account for scr refresh
        key_resp_3.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp_3, 'tStartRefresh')  # time at next scr refresh
        key_resp_3.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp_3.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp_3.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp_3.status == STARTED and not waitOnFlip:
        theseKeys = key_resp_3.getKeys(keyList=None, waitRelease=False)
        _key_resp_3_allKeys.extend(theseKeys)
        if len(_key_resp_3_allKeys):
            key_resp_3.keys = _key_resp_3_allKeys[-1].name  # just the last key pressed
            key_resp_3.rt = _key_resp_3_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in instr_train_sComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "instr_train_s"-------
for thisComponent in instr_train_sComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('text_3.started', text_3.tStartRefresh)
thisExp.addData('text_3.stopped', text_3.tStopRefresh)
# the Routine "instr_train_s" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
training_social = data.TrialHandler(nReps=1, method='random', 
    extraInfo=expInfo, originPath=-1,
    trialList=data.importConditions('trainingtrials_social.xlsx', selection='0:4'),
    seed=None, name='training_social')
thisExp.addLoop(training_social)  # add the loop to the experiment
thisTraining_social = training_social.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb = thisTraining_social.rgb)
if thisTraining_social != None:
    for paramName in thisTraining_social:
        exec('{} = thisTraining_social[paramName]'.format(paramName))

for thisTraining_social in training_social:
    currentLoop = training_social
    # abbreviate parameter names if possible (e.g. rgb = thisTraining_social.rgb)
    if thisTraining_social != None:
        for paramName in thisTraining_social:
            exec('{} = thisTraining_social[paramName]'.format(paramName))
    
    # ------Prepare to start Routine "trial"-------
    continueRoutine = True
    routineTimer.add(7.000000)
    # update component parameters for each repeat
    neutral_still.setImage(filename_neutral)
    response_training.keys = []
    response_training.rt = []
    _response_training_allKeys = []
    Stimulus.setText(str(int(Stim)))
    left_disp.setText(left_cat)
    right_disp.setText(right_cat)
    # keep track of which components have finished
    trialComponents = [fix_cross, neutral_still, response_training, Stimulus, left_disp, right_disp]
    for thisComponent in trialComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    trialClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "trial"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = trialClock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=trialClock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *fix_cross* updates
        if fix_cross.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            fix_cross.frameNStart = frameN  # exact frame index
            fix_cross.tStart = t  # local t and not account for scr refresh
            fix_cross.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(fix_cross, 'tStartRefresh')  # time at next scr refresh
            fix_cross.setAutoDraw(True)
        if fix_cross.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > fix_cross.tStartRefresh + 1-frameTolerance:
                # keep track of stop time/frame for later
                fix_cross.tStop = t  # not accounting for scr refresh
                fix_cross.frameNStop = frameN  # exact frame index
                win.timeOnFlip(fix_cross, 'tStopRefresh')  # time at next scr refresh
                fix_cross.setAutoDraw(False)
        
        # *neutral_still* updates
        if neutral_still.status == NOT_STARTED and tThisFlip >= 1.0-frameTolerance:
            # keep track of start time/frame for later
            neutral_still.frameNStart = frameN  # exact frame index
            neutral_still.tStart = t  # local t and not account for scr refresh
            neutral_still.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(neutral_still, 'tStartRefresh')  # time at next scr refresh
            neutral_still.setAutoDraw(True)
        if neutral_still.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > neutral_still.tStartRefresh + 6.0-frameTolerance:
                # keep track of stop time/frame for later
                neutral_still.tStop = t  # not accounting for scr refresh
                neutral_still.frameNStop = frameN  # exact frame index
                win.timeOnFlip(neutral_still, 'tStopRefresh')  # time at next scr refresh
                neutral_still.setAutoDraw(False)
        
        # *response_training* updates
        waitOnFlip = False
        if response_training.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            response_training.frameNStart = frameN  # exact frame index
            response_training.tStart = t  # local t and not account for scr refresh
            response_training.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(response_training, 'tStartRefresh')  # time at next scr refresh
            response_training.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(response_training.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(response_training.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if response_training.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > response_training.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                response_training.tStop = t  # not accounting for scr refresh
                response_training.frameNStop = frameN  # exact frame index
                win.timeOnFlip(response_training, 'tStopRefresh')  # time at next scr refresh
                response_training.status = FINISHED
        if response_training.status == STARTED and not waitOnFlip:
            theseKeys = response_training.getKeys(keyList=['left', 'right'], waitRelease=False)
            _response_training_allKeys.extend(theseKeys)
            if len(_response_training_allKeys):
                response_training.keys = _response_training_allKeys[-1].name  # just the last key pressed
                response_training.rt = _response_training_allKeys[-1].rt
                # was this correct?
                if (response_training.keys == str(CorrCat)) or (response_training.keys == CorrCat):
                    response_training.corr = 1
                else:
                    response_training.corr = 0
                # a response ends the routine
                continueRoutine = False
        
        # *Stimulus* updates
        if Stimulus.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            Stimulus.frameNStart = frameN  # exact frame index
            Stimulus.tStart = t  # local t and not account for scr refresh
            Stimulus.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(Stimulus, 'tStartRefresh')  # time at next scr refresh
            Stimulus.setAutoDraw(True)
        if Stimulus.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > Stimulus.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                Stimulus.tStop = t  # not accounting for scr refresh
                Stimulus.frameNStop = frameN  # exact frame index
                win.timeOnFlip(Stimulus, 'tStopRefresh')  # time at next scr refresh
                Stimulus.setAutoDraw(False)
        
        # *left_disp* updates
        if left_disp.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            left_disp.frameNStart = frameN  # exact frame index
            left_disp.tStart = t  # local t and not account for scr refresh
            left_disp.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(left_disp, 'tStartRefresh')  # time at next scr refresh
            left_disp.setAutoDraw(True)
        if left_disp.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > left_disp.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                left_disp.tStop = t  # not accounting for scr refresh
                left_disp.frameNStop = frameN  # exact frame index
                win.timeOnFlip(left_disp, 'tStopRefresh')  # time at next scr refresh
                left_disp.setAutoDraw(False)
        
        # *right_disp* updates
        if right_disp.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            right_disp.frameNStart = frameN  # exact frame index
            right_disp.tStart = t  # local t and not account for scr refresh
            right_disp.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(right_disp, 'tStartRefresh')  # time at next scr refresh
            right_disp.setAutoDraw(True)
        if right_disp.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > right_disp.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                right_disp.tStop = t  # not accounting for scr refresh
                right_disp.frameNStop = frameN  # exact frame index
                win.timeOnFlip(right_disp, 'tStopRefresh')  # time at next scr refresh
                right_disp.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in trialComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "trial"-------
    for thisComponent in trialComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    #Speichern von keyangry und keyhappy ins Data file in jedem Trial
    thisExp.addData ('left', left_cat)
    thisExp.addData ('right', right_cat)
    thisExp.addData('stim', Stim)
    
    if response_training.keys == 'left':
        thisExp.addData('ParticipantAnswer', left_cat)
    elif response_training.keys == 'right':
        thisExp.addData('ParticipantAnswer', right_cat)
    elif response_training.keys == None:
        thisExp.addData('ParticipantAnswer', None)
    training_social.addData('fix_cross.started', fix_cross.tStartRefresh)
    training_social.addData('fix_cross.stopped', fix_cross.tStopRefresh)
    training_social.addData('neutral_still.started', neutral_still.tStartRefresh)
    training_social.addData('neutral_still.stopped', neutral_still.tStopRefresh)
    # check responses
    if response_training.keys in ['', [], None]:  # No response was made
        response_training.keys = None
        # was no response the correct answer?!
        if str(CorrCat).lower() == 'none':
           response_training.corr = 1;  # correct non-response
        else:
           response_training.corr = 0;  # failed to respond (incorrectly)
    # store data for training_social (TrialHandler)
    training_social.addData('response_training.keys',response_training.keys)
    training_social.addData('response_training.corr', response_training.corr)
    if response_training.keys != None:  # we had a response
        training_social.addData('response_training.rt', response_training.rt)
    training_social.addData('response_training.started', response_training.tStartRefresh)
    training_social.addData('response_training.stopped', response_training.tStopRefresh)
    training_social.addData('Stimulus.started', Stimulus.tStartRefresh)
    training_social.addData('Stimulus.stopped', Stimulus.tStopRefresh)
    training_social.addData('left_disp.started', left_disp.tStartRefresh)
    training_social.addData('left_disp.stopped', left_disp.tStopRefresh)
    training_social.addData('right_disp.started', right_disp.tStartRefresh)
    training_social.addData('right_disp.stopped', right_disp.tStopRefresh)
    
    # ------Prepare to start Routine "feedback_train"-------
    continueRoutine = True
    # update component parameters for each repeat
    #Feedback Training
    
    #Show happy/angry in 50% of cases
    if CorrCat == left_cat:
        time_h = 6.0
        time_a = 0.0
    else:
        time_h = 0.0
        time_a = 6.0
    
    #Set color of A/B on screen after response (and feedback for miss)
    time_miss = 0
    if response_training.keys == 'left':
        left_color = 'black'
        right_color = 'grey'
    elif response_training.keys == 'right':
        left_color = 'grey'
        right_color = 'black'
    elif response_training.keys == None:
        left_color = 'grey'
        right_color = 'grey'
        time_miss = 6.0
        time_h = 0.0
        time_a = 0.0
    new_angry_2 = visual.MovieStim3(
        win=win, name='new_angry_2',units='height', 
        noAudio = True,
        filename=filename_feedback_angry,
        ori=0, pos=[0, 0.225], opacity=1,
        loop=False,
        size=[0.5, 0.5],
        depth=-2.0,
        )
    new_happy_2 = visual.MovieStim3(
        win=win, name='new_happy_2',units='height', 
        noAudio = False,
        filename=filename_feedback_happy,
        ori=0, pos=[0, 0.225], opacity=1,
        loop=False,
        size=[0.5, 0.5],
        depth=-3.0,
        )
    left_disp_fb_2.setColor(left_color, colorSpace='rgb')
    left_disp_fb_2.setText(left_cat)
    right_disp_fb_2.setColor(right_color, colorSpace='rgb')
    right_disp_fb_2.setText(right_cat)
    Stimulus_fb_2.setText(str(int(Stim)))
    # keep track of which components have finished
    feedback_trainComponents = [feedback_miss, new_angry_2, new_happy_2, left_disp_fb_2, right_disp_fb_2, Stimulus_fb_2]
    for thisComponent in feedback_trainComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    feedback_trainClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "feedback_train"-------
    while continueRoutine:
        # get current time
        t = feedback_trainClock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=feedback_trainClock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *feedback_miss* updates
        if feedback_miss.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            feedback_miss.frameNStart = frameN  # exact frame index
            feedback_miss.tStart = t  # local t and not account for scr refresh
            feedback_miss.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(feedback_miss, 'tStartRefresh')  # time at next scr refresh
            feedback_miss.setAutoDraw(True)
        if feedback_miss.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > feedback_miss.tStartRefresh + time_miss-frameTolerance:
                # keep track of stop time/frame for later
                feedback_miss.tStop = t  # not accounting for scr refresh
                feedback_miss.frameNStop = frameN  # exact frame index
                win.timeOnFlip(feedback_miss, 'tStopRefresh')  # time at next scr refresh
                feedback_miss.setAutoDraw(False)
        
        # *new_angry_2* updates
        if new_angry_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            new_angry_2.frameNStart = frameN  # exact frame index
            new_angry_2.tStart = t  # local t and not account for scr refresh
            new_angry_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(new_angry_2, 'tStartRefresh')  # time at next scr refresh
            new_angry_2.setAutoDraw(True)
        if new_angry_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > new_angry_2.tStartRefresh + time_a-frameTolerance:
                # keep track of stop time/frame for later
                new_angry_2.tStop = t  # not accounting for scr refresh
                new_angry_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(new_angry_2, 'tStopRefresh')  # time at next scr refresh
                new_angry_2.setAutoDraw(False)
        
        # *new_happy_2* updates
        if new_happy_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            new_happy_2.frameNStart = frameN  # exact frame index
            new_happy_2.tStart = t  # local t and not account for scr refresh
            new_happy_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(new_happy_2, 'tStartRefresh')  # time at next scr refresh
            new_happy_2.setAutoDraw(True)
        if new_happy_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > new_happy_2.tStartRefresh + time_h-frameTolerance:
                # keep track of stop time/frame for later
                new_happy_2.tStop = t  # not accounting for scr refresh
                new_happy_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(new_happy_2, 'tStopRefresh')  # time at next scr refresh
                new_happy_2.setAutoDraw(False)
        
        # *left_disp_fb_2* updates
        if left_disp_fb_2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            left_disp_fb_2.frameNStart = frameN  # exact frame index
            left_disp_fb_2.tStart = t  # local t and not account for scr refresh
            left_disp_fb_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(left_disp_fb_2, 'tStartRefresh')  # time at next scr refresh
            left_disp_fb_2.setAutoDraw(True)
        if left_disp_fb_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > left_disp_fb_2.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                left_disp_fb_2.tStop = t  # not accounting for scr refresh
                left_disp_fb_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(left_disp_fb_2, 'tStopRefresh')  # time at next scr refresh
                left_disp_fb_2.setAutoDraw(False)
        
        # *right_disp_fb_2* updates
        if right_disp_fb_2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            right_disp_fb_2.frameNStart = frameN  # exact frame index
            right_disp_fb_2.tStart = t  # local t and not account for scr refresh
            right_disp_fb_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(right_disp_fb_2, 'tStartRefresh')  # time at next scr refresh
            right_disp_fb_2.setAutoDraw(True)
        if right_disp_fb_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > right_disp_fb_2.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                right_disp_fb_2.tStop = t  # not accounting for scr refresh
                right_disp_fb_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(right_disp_fb_2, 'tStopRefresh')  # time at next scr refresh
                right_disp_fb_2.setAutoDraw(False)
        
        # *Stimulus_fb_2* updates
        if Stimulus_fb_2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            Stimulus_fb_2.frameNStart = frameN  # exact frame index
            Stimulus_fb_2.tStart = t  # local t and not account for scr refresh
            Stimulus_fb_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(Stimulus_fb_2, 'tStartRefresh')  # time at next scr refresh
            Stimulus_fb_2.setAutoDraw(True)
        if Stimulus_fb_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > Stimulus_fb_2.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                Stimulus_fb_2.tStop = t  # not accounting for scr refresh
                Stimulus_fb_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(Stimulus_fb_2, 'tStopRefresh')  # time at next scr refresh
                Stimulus_fb_2.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in feedback_trainComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "feedback_train"-------
    for thisComponent in feedback_trainComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    training_social.addData('feedback_miss.started', feedback_miss.tStartRefresh)
    training_social.addData('feedback_miss.stopped', feedback_miss.tStopRefresh)
    new_angry_2.stop()
    new_happy_2.stop()
    training_social.addData('left_disp_fb_2.started', left_disp_fb_2.tStartRefresh)
    training_social.addData('left_disp_fb_2.stopped', left_disp_fb_2.tStopRefresh)
    training_social.addData('right_disp_fb_2.started', right_disp_fb_2.tStartRefresh)
    training_social.addData('right_disp_fb_2.stopped', right_disp_fb_2.tStopRefresh)
    training_social.addData('Stimulus_fb_2.started', Stimulus_fb_2.tStartRefresh)
    training_social.addData('Stimulus_fb_2.stopped', Stimulus_fb_2.tStopRefresh)
    # the Routine "feedback_train" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    thisExp.nextEntry()
    
# completed 1 repeats of 'training_social'


# ------Prepare to start Routine "Alles_klar"-------
continueRoutine = True
# update component parameters for each repeat
key_resp_4.keys = []
key_resp_4.rt = []
_key_resp_4_allKeys = []
# keep track of which components have finished
Alles_klarComponents = [text_4, key_resp_4]
for thisComponent in Alles_klarComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
Alles_klarClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "Alles_klar"-------
while continueRoutine:
    # get current time
    t = Alles_klarClock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=Alles_klarClock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *text_4* updates
    if text_4.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        text_4.frameNStart = frameN  # exact frame index
        text_4.tStart = t  # local t and not account for scr refresh
        text_4.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(text_4, 'tStartRefresh')  # time at next scr refresh
        text_4.setAutoDraw(True)
    
    # *key_resp_4* updates
    waitOnFlip = False
    if key_resp_4.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp_4.frameNStart = frameN  # exact frame index
        key_resp_4.tStart = t  # local t and not account for scr refresh
        key_resp_4.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp_4, 'tStartRefresh')  # time at next scr refresh
        key_resp_4.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp_4.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp_4.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp_4.status == STARTED and not waitOnFlip:
        theseKeys = key_resp_4.getKeys(keyList=None, waitRelease=False)
        _key_resp_4_allKeys.extend(theseKeys)
        if len(_key_resp_4_allKeys):
            key_resp_4.keys = _key_resp_4_allKeys[-1].name  # just the last key pressed
            key_resp_4.rt = _key_resp_4_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in Alles_klarComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "Alles_klar"-------
for thisComponent in Alles_klarComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('text_4.started', text_4.tStartRefresh)
thisExp.addData('text_4.stopped', text_4.tStopRefresh)
# the Routine "Alles_klar" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# ------Prepare to start Routine "instr_train_ns"-------
continueRoutine = True
# update component parameters for each repeat
key_resp_5.keys = []
key_resp_5.rt = []
_key_resp_5_allKeys = []
# keep track of which components have finished
instr_train_nsComponents = [text, key_resp_5]
for thisComponent in instr_train_nsComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
instr_train_nsClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "instr_train_ns"-------
while continueRoutine:
    # get current time
    t = instr_train_nsClock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=instr_train_nsClock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *text* updates
    if text.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        text.frameNStart = frameN  # exact frame index
        text.tStart = t  # local t and not account for scr refresh
        text.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(text, 'tStartRefresh')  # time at next scr refresh
        text.setAutoDraw(True)
    
    # *key_resp_5* updates
    waitOnFlip = False
    if key_resp_5.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        key_resp_5.frameNStart = frameN  # exact frame index
        key_resp_5.tStart = t  # local t and not account for scr refresh
        key_resp_5.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(key_resp_5, 'tStartRefresh')  # time at next scr refresh
        key_resp_5.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(key_resp_5.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(key_resp_5.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if key_resp_5.status == STARTED and not waitOnFlip:
        theseKeys = key_resp_5.getKeys(keyList=None, waitRelease=False)
        _key_resp_5_allKeys.extend(theseKeys)
        if len(_key_resp_5_allKeys):
            key_resp_5.keys = _key_resp_5_allKeys[-1].name  # just the last key pressed
            key_resp_5.rt = _key_resp_5_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in instr_train_nsComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "instr_train_ns"-------
for thisComponent in instr_train_nsComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('text.started', text.tStartRefresh)
thisExp.addData('text.stopped', text.tStopRefresh)
# the Routine "instr_train_ns" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
training_nonsocial = data.TrialHandler(nReps=1, method='random', 
    extraInfo=expInfo, originPath=-1,
    trialList=data.importConditions('trainingtrials_nonsocial.xlsx', selection='0:4'),
    seed=None, name='training_nonsocial')
thisExp.addLoop(training_nonsocial)  # add the loop to the experiment
thisTraining_nonsocial = training_nonsocial.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb = thisTraining_nonsocial.rgb)
if thisTraining_nonsocial != None:
    for paramName in thisTraining_nonsocial:
        exec('{} = thisTraining_nonsocial[paramName]'.format(paramName))

for thisTraining_nonsocial in training_nonsocial:
    currentLoop = training_nonsocial
    # abbreviate parameter names if possible (e.g. rgb = thisTraining_nonsocial.rgb)
    if thisTraining_nonsocial != None:
        for paramName in thisTraining_nonsocial:
            exec('{} = thisTraining_nonsocial[paramName]'.format(paramName))
    
    # ------Prepare to start Routine "trial"-------
    continueRoutine = True
    routineTimer.add(7.000000)
    # update component parameters for each repeat
    neutral_still.setImage(filename_neutral)
    response_training.keys = []
    response_training.rt = []
    _response_training_allKeys = []
    Stimulus.setText(str(int(Stim)))
    left_disp.setText(left_cat)
    right_disp.setText(right_cat)
    # keep track of which components have finished
    trialComponents = [fix_cross, neutral_still, response_training, Stimulus, left_disp, right_disp]
    for thisComponent in trialComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    trialClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "trial"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = trialClock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=trialClock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *fix_cross* updates
        if fix_cross.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            fix_cross.frameNStart = frameN  # exact frame index
            fix_cross.tStart = t  # local t and not account for scr refresh
            fix_cross.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(fix_cross, 'tStartRefresh')  # time at next scr refresh
            fix_cross.setAutoDraw(True)
        if fix_cross.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > fix_cross.tStartRefresh + 1-frameTolerance:
                # keep track of stop time/frame for later
                fix_cross.tStop = t  # not accounting for scr refresh
                fix_cross.frameNStop = frameN  # exact frame index
                win.timeOnFlip(fix_cross, 'tStopRefresh')  # time at next scr refresh
                fix_cross.setAutoDraw(False)
        
        # *neutral_still* updates
        if neutral_still.status == NOT_STARTED and tThisFlip >= 1.0-frameTolerance:
            # keep track of start time/frame for later
            neutral_still.frameNStart = frameN  # exact frame index
            neutral_still.tStart = t  # local t and not account for scr refresh
            neutral_still.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(neutral_still, 'tStartRefresh')  # time at next scr refresh
            neutral_still.setAutoDraw(True)
        if neutral_still.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > neutral_still.tStartRefresh + 6.0-frameTolerance:
                # keep track of stop time/frame for later
                neutral_still.tStop = t  # not accounting for scr refresh
                neutral_still.frameNStop = frameN  # exact frame index
                win.timeOnFlip(neutral_still, 'tStopRefresh')  # time at next scr refresh
                neutral_still.setAutoDraw(False)
        
        # *response_training* updates
        waitOnFlip = False
        if response_training.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            response_training.frameNStart = frameN  # exact frame index
            response_training.tStart = t  # local t and not account for scr refresh
            response_training.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(response_training, 'tStartRefresh')  # time at next scr refresh
            response_training.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(response_training.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(response_training.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if response_training.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > response_training.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                response_training.tStop = t  # not accounting for scr refresh
                response_training.frameNStop = frameN  # exact frame index
                win.timeOnFlip(response_training, 'tStopRefresh')  # time at next scr refresh
                response_training.status = FINISHED
        if response_training.status == STARTED and not waitOnFlip:
            theseKeys = response_training.getKeys(keyList=['left', 'right'], waitRelease=False)
            _response_training_allKeys.extend(theseKeys)
            if len(_response_training_allKeys):
                response_training.keys = _response_training_allKeys[-1].name  # just the last key pressed
                response_training.rt = _response_training_allKeys[-1].rt
                # was this correct?
                if (response_training.keys == str(CorrCat)) or (response_training.keys == CorrCat):
                    response_training.corr = 1
                else:
                    response_training.corr = 0
                # a response ends the routine
                continueRoutine = False
        
        # *Stimulus* updates
        if Stimulus.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            Stimulus.frameNStart = frameN  # exact frame index
            Stimulus.tStart = t  # local t and not account for scr refresh
            Stimulus.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(Stimulus, 'tStartRefresh')  # time at next scr refresh
            Stimulus.setAutoDraw(True)
        if Stimulus.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > Stimulus.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                Stimulus.tStop = t  # not accounting for scr refresh
                Stimulus.frameNStop = frameN  # exact frame index
                win.timeOnFlip(Stimulus, 'tStopRefresh')  # time at next scr refresh
                Stimulus.setAutoDraw(False)
        
        # *left_disp* updates
        if left_disp.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            left_disp.frameNStart = frameN  # exact frame index
            left_disp.tStart = t  # local t and not account for scr refresh
            left_disp.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(left_disp, 'tStartRefresh')  # time at next scr refresh
            left_disp.setAutoDraw(True)
        if left_disp.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > left_disp.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                left_disp.tStop = t  # not accounting for scr refresh
                left_disp.frameNStop = frameN  # exact frame index
                win.timeOnFlip(left_disp, 'tStopRefresh')  # time at next scr refresh
                left_disp.setAutoDraw(False)
        
        # *right_disp* updates
        if right_disp.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            right_disp.frameNStart = frameN  # exact frame index
            right_disp.tStart = t  # local t and not account for scr refresh
            right_disp.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(right_disp, 'tStartRefresh')  # time at next scr refresh
            right_disp.setAutoDraw(True)
        if right_disp.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > right_disp.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                right_disp.tStop = t  # not accounting for scr refresh
                right_disp.frameNStop = frameN  # exact frame index
                win.timeOnFlip(right_disp, 'tStopRefresh')  # time at next scr refresh
                right_disp.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in trialComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "trial"-------
    for thisComponent in trialComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    #Speichern von keyangry und keyhappy ins Data file in jedem Trial
    thisExp.addData ('left', left_cat)
    thisExp.addData ('right', right_cat)
    thisExp.addData('stim', Stim)
    
    if response_training.keys == 'left':
        thisExp.addData('ParticipantAnswer', left_cat)
    elif response_training.keys == 'right':
        thisExp.addData('ParticipantAnswer', right_cat)
    elif response_training.keys == None:
        thisExp.addData('ParticipantAnswer', None)
    training_nonsocial.addData('fix_cross.started', fix_cross.tStartRefresh)
    training_nonsocial.addData('fix_cross.stopped', fix_cross.tStopRefresh)
    training_nonsocial.addData('neutral_still.started', neutral_still.tStartRefresh)
    training_nonsocial.addData('neutral_still.stopped', neutral_still.tStopRefresh)
    # check responses
    if response_training.keys in ['', [], None]:  # No response was made
        response_training.keys = None
        # was no response the correct answer?!
        if str(CorrCat).lower() == 'none':
           response_training.corr = 1;  # correct non-response
        else:
           response_training.corr = 0;  # failed to respond (incorrectly)
    # store data for training_nonsocial (TrialHandler)
    training_nonsocial.addData('response_training.keys',response_training.keys)
    training_nonsocial.addData('response_training.corr', response_training.corr)
    if response_training.keys != None:  # we had a response
        training_nonsocial.addData('response_training.rt', response_training.rt)
    training_nonsocial.addData('response_training.started', response_training.tStartRefresh)
    training_nonsocial.addData('response_training.stopped', response_training.tStopRefresh)
    training_nonsocial.addData('Stimulus.started', Stimulus.tStartRefresh)
    training_nonsocial.addData('Stimulus.stopped', Stimulus.tStopRefresh)
    training_nonsocial.addData('left_disp.started', left_disp.tStartRefresh)
    training_nonsocial.addData('left_disp.stopped', left_disp.tStopRefresh)
    training_nonsocial.addData('right_disp.started', right_disp.tStartRefresh)
    training_nonsocial.addData('right_disp.stopped', right_disp.tStopRefresh)
    
    # ------Prepare to start Routine "feedback_train"-------
    continueRoutine = True
    # update component parameters for each repeat
    #Feedback Training
    
    #Show happy/angry in 50% of cases
    if CorrCat == left_cat:
        time_h = 6.0
        time_a = 0.0
    else:
        time_h = 0.0
        time_a = 6.0
    
    #Set color of A/B on screen after response (and feedback for miss)
    time_miss = 0
    if response_training.keys == 'left':
        left_color = 'black'
        right_color = 'grey'
    elif response_training.keys == 'right':
        left_color = 'grey'
        right_color = 'black'
    elif response_training.keys == None:
        left_color = 'grey'
        right_color = 'grey'
        time_miss = 6.0
        time_h = 0.0
        time_a = 0.0
    new_angry_2 = visual.MovieStim3(
        win=win, name='new_angry_2',units='height', 
        noAudio = True,
        filename=filename_feedback_angry,
        ori=0, pos=[0, 0.225], opacity=1,
        loop=False,
        size=[0.5, 0.5],
        depth=-2.0,
        )
    new_happy_2 = visual.MovieStim3(
        win=win, name='new_happy_2',units='height', 
        noAudio = False,
        filename=filename_feedback_happy,
        ori=0, pos=[0, 0.225], opacity=1,
        loop=False,
        size=[0.5, 0.5],
        depth=-3.0,
        )
    left_disp_fb_2.setColor(left_color, colorSpace='rgb')
    left_disp_fb_2.setText(left_cat)
    right_disp_fb_2.setColor(right_color, colorSpace='rgb')
    right_disp_fb_2.setText(right_cat)
    Stimulus_fb_2.setText(str(int(Stim)))
    # keep track of which components have finished
    feedback_trainComponents = [feedback_miss, new_angry_2, new_happy_2, left_disp_fb_2, right_disp_fb_2, Stimulus_fb_2]
    for thisComponent in feedback_trainComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    feedback_trainClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "feedback_train"-------
    while continueRoutine:
        # get current time
        t = feedback_trainClock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=feedback_trainClock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *feedback_miss* updates
        if feedback_miss.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            feedback_miss.frameNStart = frameN  # exact frame index
            feedback_miss.tStart = t  # local t and not account for scr refresh
            feedback_miss.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(feedback_miss, 'tStartRefresh')  # time at next scr refresh
            feedback_miss.setAutoDraw(True)
        if feedback_miss.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > feedback_miss.tStartRefresh + time_miss-frameTolerance:
                # keep track of stop time/frame for later
                feedback_miss.tStop = t  # not accounting for scr refresh
                feedback_miss.frameNStop = frameN  # exact frame index
                win.timeOnFlip(feedback_miss, 'tStopRefresh')  # time at next scr refresh
                feedback_miss.setAutoDraw(False)
        
        # *new_angry_2* updates
        if new_angry_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            new_angry_2.frameNStart = frameN  # exact frame index
            new_angry_2.tStart = t  # local t and not account for scr refresh
            new_angry_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(new_angry_2, 'tStartRefresh')  # time at next scr refresh
            new_angry_2.setAutoDraw(True)
        if new_angry_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > new_angry_2.tStartRefresh + time_a-frameTolerance:
                # keep track of stop time/frame for later
                new_angry_2.tStop = t  # not accounting for scr refresh
                new_angry_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(new_angry_2, 'tStopRefresh')  # time at next scr refresh
                new_angry_2.setAutoDraw(False)
        
        # *new_happy_2* updates
        if new_happy_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            new_happy_2.frameNStart = frameN  # exact frame index
            new_happy_2.tStart = t  # local t and not account for scr refresh
            new_happy_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(new_happy_2, 'tStartRefresh')  # time at next scr refresh
            new_happy_2.setAutoDraw(True)
        if new_happy_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > new_happy_2.tStartRefresh + time_h-frameTolerance:
                # keep track of stop time/frame for later
                new_happy_2.tStop = t  # not accounting for scr refresh
                new_happy_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(new_happy_2, 'tStopRefresh')  # time at next scr refresh
                new_happy_2.setAutoDraw(False)
        
        # *left_disp_fb_2* updates
        if left_disp_fb_2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            left_disp_fb_2.frameNStart = frameN  # exact frame index
            left_disp_fb_2.tStart = t  # local t and not account for scr refresh
            left_disp_fb_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(left_disp_fb_2, 'tStartRefresh')  # time at next scr refresh
            left_disp_fb_2.setAutoDraw(True)
        if left_disp_fb_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > left_disp_fb_2.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                left_disp_fb_2.tStop = t  # not accounting for scr refresh
                left_disp_fb_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(left_disp_fb_2, 'tStopRefresh')  # time at next scr refresh
                left_disp_fb_2.setAutoDraw(False)
        
        # *right_disp_fb_2* updates
        if right_disp_fb_2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            right_disp_fb_2.frameNStart = frameN  # exact frame index
            right_disp_fb_2.tStart = t  # local t and not account for scr refresh
            right_disp_fb_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(right_disp_fb_2, 'tStartRefresh')  # time at next scr refresh
            right_disp_fb_2.setAutoDraw(True)
        if right_disp_fb_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > right_disp_fb_2.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                right_disp_fb_2.tStop = t  # not accounting for scr refresh
                right_disp_fb_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(right_disp_fb_2, 'tStopRefresh')  # time at next scr refresh
                right_disp_fb_2.setAutoDraw(False)
        
        # *Stimulus_fb_2* updates
        if Stimulus_fb_2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            Stimulus_fb_2.frameNStart = frameN  # exact frame index
            Stimulus_fb_2.tStart = t  # local t and not account for scr refresh
            Stimulus_fb_2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(Stimulus_fb_2, 'tStartRefresh')  # time at next scr refresh
            Stimulus_fb_2.setAutoDraw(True)
        if Stimulus_fb_2.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > Stimulus_fb_2.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                Stimulus_fb_2.tStop = t  # not accounting for scr refresh
                Stimulus_fb_2.frameNStop = frameN  # exact frame index
                win.timeOnFlip(Stimulus_fb_2, 'tStopRefresh')  # time at next scr refresh
                Stimulus_fb_2.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in feedback_trainComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "feedback_train"-------
    for thisComponent in feedback_trainComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    training_nonsocial.addData('feedback_miss.started', feedback_miss.tStartRefresh)
    training_nonsocial.addData('feedback_miss.stopped', feedback_miss.tStopRefresh)
    new_angry_2.stop()
    new_happy_2.stop()
    training_nonsocial.addData('left_disp_fb_2.started', left_disp_fb_2.tStartRefresh)
    training_nonsocial.addData('left_disp_fb_2.stopped', left_disp_fb_2.tStopRefresh)
    training_nonsocial.addData('right_disp_fb_2.started', right_disp_fb_2.tStartRefresh)
    training_nonsocial.addData('right_disp_fb_2.stopped', right_disp_fb_2.tStopRefresh)
    training_nonsocial.addData('Stimulus_fb_2.started', Stimulus_fb_2.tStartRefresh)
    training_nonsocial.addData('Stimulus_fb_2.stopped', Stimulus_fb_2.tStopRefresh)
    # the Routine "feedback_train" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    thisExp.nextEntry()
    
# completed 1 repeats of 'training_nonsocial'


# ------Prepare to start Routine "Alles_Klar_2"-------
continueRoutine = True
# update component parameters for each repeat
Allesklar_resp_2.keys = []
Allesklar_resp_2.rt = []
_Allesklar_resp_2_allKeys = []
# keep track of which components have finished
Alles_Klar_2Components = [Allesklartext_2, Allesklar_resp_2]
for thisComponent in Alles_Klar_2Components:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
Alles_Klar_2Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "Alles_Klar_2"-------
while continueRoutine:
    # get current time
    t = Alles_Klar_2Clock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=Alles_Klar_2Clock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *Allesklartext_2* updates
    if Allesklartext_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        Allesklartext_2.frameNStart = frameN  # exact frame index
        Allesklartext_2.tStart = t  # local t and not account for scr refresh
        Allesklartext_2.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(Allesklartext_2, 'tStartRefresh')  # time at next scr refresh
        Allesklartext_2.setAutoDraw(True)
    
    # *Allesklar_resp_2* updates
    waitOnFlip = False
    if Allesklar_resp_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        Allesklar_resp_2.frameNStart = frameN  # exact frame index
        Allesklar_resp_2.tStart = t  # local t and not account for scr refresh
        Allesklar_resp_2.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(Allesklar_resp_2, 'tStartRefresh')  # time at next scr refresh
        Allesklar_resp_2.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(Allesklar_resp_2.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(Allesklar_resp_2.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if Allesklar_resp_2.status == STARTED and not waitOnFlip:
        theseKeys = Allesklar_resp_2.getKeys(keyList=None, waitRelease=False)
        _Allesklar_resp_2_allKeys.extend(theseKeys)
        if len(_Allesklar_resp_2_allKeys):
            Allesklar_resp_2.keys = _Allesklar_resp_2_allKeys[-1].name  # just the last key pressed
            Allesklar_resp_2.rt = _Allesklar_resp_2_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in Alles_Klar_2Components:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "Alles_Klar_2"-------
for thisComponent in Alles_Klar_2Components:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('Allesklartext_2.started', Allesklartext_2.tStartRefresh)
thisExp.addData('Allesklartext_2.stopped', Allesklartext_2.tStopRefresh)
# the Routine "Alles_Klar_2" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
prob_fb = data.TrialHandler(nReps=1, method='random', 
    extraInfo=expInfo, originPath=-1,
    trialList=data.importConditions('trainingtrials_social_prob.xlsx', selection='0:4'),
    seed=None, name='prob_fb')
thisExp.addLoop(prob_fb)  # add the loop to the experiment
thisProb_fb = prob_fb.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb = thisProb_fb.rgb)
if thisProb_fb != None:
    for paramName in thisProb_fb:
        exec('{} = thisProb_fb[paramName]'.format(paramName))

for thisProb_fb in prob_fb:
    currentLoop = prob_fb
    # abbreviate parameter names if possible (e.g. rgb = thisProb_fb.rgb)
    if thisProb_fb != None:
        for paramName in thisProb_fb:
            exec('{} = thisProb_fb[paramName]'.format(paramName))
    
    # ------Prepare to start Routine "trial"-------
    continueRoutine = True
    routineTimer.add(7.000000)
    # update component parameters for each repeat
    neutral_still.setImage(filename_neutral)
    response_training.keys = []
    response_training.rt = []
    _response_training_allKeys = []
    Stimulus.setText(str(int(Stim)))
    left_disp.setText(left_cat)
    right_disp.setText(right_cat)
    # keep track of which components have finished
    trialComponents = [fix_cross, neutral_still, response_training, Stimulus, left_disp, right_disp]
    for thisComponent in trialComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    trialClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "trial"-------
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = trialClock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=trialClock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *fix_cross* updates
        if fix_cross.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            fix_cross.frameNStart = frameN  # exact frame index
            fix_cross.tStart = t  # local t and not account for scr refresh
            fix_cross.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(fix_cross, 'tStartRefresh')  # time at next scr refresh
            fix_cross.setAutoDraw(True)
        if fix_cross.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > fix_cross.tStartRefresh + 1-frameTolerance:
                # keep track of stop time/frame for later
                fix_cross.tStop = t  # not accounting for scr refresh
                fix_cross.frameNStop = frameN  # exact frame index
                win.timeOnFlip(fix_cross, 'tStopRefresh')  # time at next scr refresh
                fix_cross.setAutoDraw(False)
        
        # *neutral_still* updates
        if neutral_still.status == NOT_STARTED and tThisFlip >= 1.0-frameTolerance:
            # keep track of start time/frame for later
            neutral_still.frameNStart = frameN  # exact frame index
            neutral_still.tStart = t  # local t and not account for scr refresh
            neutral_still.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(neutral_still, 'tStartRefresh')  # time at next scr refresh
            neutral_still.setAutoDraw(True)
        if neutral_still.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > neutral_still.tStartRefresh + 6.0-frameTolerance:
                # keep track of stop time/frame for later
                neutral_still.tStop = t  # not accounting for scr refresh
                neutral_still.frameNStop = frameN  # exact frame index
                win.timeOnFlip(neutral_still, 'tStopRefresh')  # time at next scr refresh
                neutral_still.setAutoDraw(False)
        
        # *response_training* updates
        waitOnFlip = False
        if response_training.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            response_training.frameNStart = frameN  # exact frame index
            response_training.tStart = t  # local t and not account for scr refresh
            response_training.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(response_training, 'tStartRefresh')  # time at next scr refresh
            response_training.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(response_training.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(response_training.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if response_training.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > response_training.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                response_training.tStop = t  # not accounting for scr refresh
                response_training.frameNStop = frameN  # exact frame index
                win.timeOnFlip(response_training, 'tStopRefresh')  # time at next scr refresh
                response_training.status = FINISHED
        if response_training.status == STARTED and not waitOnFlip:
            theseKeys = response_training.getKeys(keyList=['left', 'right'], waitRelease=False)
            _response_training_allKeys.extend(theseKeys)
            if len(_response_training_allKeys):
                response_training.keys = _response_training_allKeys[-1].name  # just the last key pressed
                response_training.rt = _response_training_allKeys[-1].rt
                # was this correct?
                if (response_training.keys == str(CorrCat)) or (response_training.keys == CorrCat):
                    response_training.corr = 1
                else:
                    response_training.corr = 0
                # a response ends the routine
                continueRoutine = False
        
        # *Stimulus* updates
        if Stimulus.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            Stimulus.frameNStart = frameN  # exact frame index
            Stimulus.tStart = t  # local t and not account for scr refresh
            Stimulus.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(Stimulus, 'tStartRefresh')  # time at next scr refresh
            Stimulus.setAutoDraw(True)
        if Stimulus.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > Stimulus.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                Stimulus.tStop = t  # not accounting for scr refresh
                Stimulus.frameNStop = frameN  # exact frame index
                win.timeOnFlip(Stimulus, 'tStopRefresh')  # time at next scr refresh
                Stimulus.setAutoDraw(False)
        
        # *left_disp* updates
        if left_disp.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            left_disp.frameNStart = frameN  # exact frame index
            left_disp.tStart = t  # local t and not account for scr refresh
            left_disp.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(left_disp, 'tStartRefresh')  # time at next scr refresh
            left_disp.setAutoDraw(True)
        if left_disp.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > left_disp.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                left_disp.tStop = t  # not accounting for scr refresh
                left_disp.frameNStop = frameN  # exact frame index
                win.timeOnFlip(left_disp, 'tStopRefresh')  # time at next scr refresh
                left_disp.setAutoDraw(False)
        
        # *right_disp* updates
        if right_disp.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
            # keep track of start time/frame for later
            right_disp.frameNStart = frameN  # exact frame index
            right_disp.tStart = t  # local t and not account for scr refresh
            right_disp.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(right_disp, 'tStartRefresh')  # time at next scr refresh
            right_disp.setAutoDraw(True)
        if right_disp.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > right_disp.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                right_disp.tStop = t  # not accounting for scr refresh
                right_disp.frameNStop = frameN  # exact frame index
                win.timeOnFlip(right_disp, 'tStopRefresh')  # time at next scr refresh
                right_disp.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in trialComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "trial"-------
    for thisComponent in trialComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    #Speichern von keyangry und keyhappy ins Data file in jedem Trial
    thisExp.addData ('left', left_cat)
    thisExp.addData ('right', right_cat)
    thisExp.addData('stim', Stim)
    
    if response_training.keys == 'left':
        thisExp.addData('ParticipantAnswer', left_cat)
    elif response_training.keys == 'right':
        thisExp.addData('ParticipantAnswer', right_cat)
    elif response_training.keys == None:
        thisExp.addData('ParticipantAnswer', None)
    prob_fb.addData('fix_cross.started', fix_cross.tStartRefresh)
    prob_fb.addData('fix_cross.stopped', fix_cross.tStopRefresh)
    prob_fb.addData('neutral_still.started', neutral_still.tStartRefresh)
    prob_fb.addData('neutral_still.stopped', neutral_still.tStopRefresh)
    # check responses
    if response_training.keys in ['', [], None]:  # No response was made
        response_training.keys = None
        # was no response the correct answer?!
        if str(CorrCat).lower() == 'none':
           response_training.corr = 1;  # correct non-response
        else:
           response_training.corr = 0;  # failed to respond (incorrectly)
    # store data for prob_fb (TrialHandler)
    prob_fb.addData('response_training.keys',response_training.keys)
    prob_fb.addData('response_training.corr', response_training.corr)
    if response_training.keys != None:  # we had a response
        prob_fb.addData('response_training.rt', response_training.rt)
    prob_fb.addData('response_training.started', response_training.tStartRefresh)
    prob_fb.addData('response_training.stopped', response_training.tStopRefresh)
    prob_fb.addData('Stimulus.started', Stimulus.tStartRefresh)
    prob_fb.addData('Stimulus.stopped', Stimulus.tStopRefresh)
    prob_fb.addData('left_disp.started', left_disp.tStartRefresh)
    prob_fb.addData('left_disp.stopped', left_disp.tStopRefresh)
    prob_fb.addData('right_disp.started', right_disp.tStartRefresh)
    prob_fb.addData('right_disp.stopped', right_disp.tStopRefresh)
    
    # ------Prepare to start Routine "feedback_prob"-------
    continueRoutine = True
    # update component parameters for each repeat
    #Feedback Zuweisung
    # to make feedback probabilistic: uncomment nested if clause
    # this will show happy for correct (congruent) in 80% of cases 
    # and angry for correct (incongruent) in 20% of cases
    # probability can be changed by manipulating the sequence in p
    
    
    #if correct show happy video (with 0.8 probability)
    if (response_training.keys == 'left' and CorrCat == left_cat) or (response_training.keys == 'right' and CorrCat == right_cat):
        prob_msg = "Richtige Antwort!"
        if choice([True, False], p = [0.8, 0.2]):
            time_h = 6.0
            time_a = 0.0
        else:
            time_h = 0.0
            time_a = 6.0
    
    #if not correct show angry video (with 0.8 probability)
    else:
        prob_msg = "Falsche Antwort!"
        if choice([True, False], p = [0.8, 0.2]):
            time_h = 0.0
            time_a = 6.0
        else:
            time_h = 6.0
            time_a = 0.0
    
    #Set color of A/B on screen after response
    time_miss = 0
    if response_training.keys == 'left':
        left_color = 'black'
        right_color = 'grey'
    elif response_training.keys == 'right':
        left_color = 'grey'
        right_color = 'black'
    elif response_training.keys == None:
        left_color = 'grey'
        right_color = 'grey'
        time_miss = 6
        time_a = 0
        time_h = 0
    new_angry_4 = visual.MovieStim3(
        win=win, name='new_angry_4',units='height', 
        noAudio = True,
        filename=filename_feedback_angry,
        ori=0, pos=[0, 0.225], opacity=1,
        loop=False,
        size=[0.5, 0.5],
        depth=-2.0,
        )
    new_happy_4 = visual.MovieStim3(
        win=win, name='new_happy_4',units='height', 
        noAudio = False,
        filename=filename_feedback_happy,
        ori=0, pos=[0, 0.225], opacity=1,
        loop=False,
        size=[0.5, 0.5],
        depth=-3.0,
        )
    left_disp_fb_4.setColor(left_color, colorSpace='rgb')
    left_disp_fb_4.setText(left_cat)
    right_disp_fb_4.setColor(right_color, colorSpace='rgb')
    right_disp_fb_4.setText(right_cat)
    Stimulus_fb_4.setText(str(int(Stim)))
    Prob_feedback.setText(prob_msg)
    # keep track of which components have finished
    feedback_probComponents = [feedback_miss2_3, new_angry_4, new_happy_4, left_disp_fb_4, right_disp_fb_4, Stimulus_fb_4, Prob_feedback]
    for thisComponent in feedback_probComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    feedback_probClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "feedback_prob"-------
    while continueRoutine:
        # get current time
        t = feedback_probClock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=feedback_probClock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *feedback_miss2_3* updates
        if feedback_miss2_3.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            feedback_miss2_3.frameNStart = frameN  # exact frame index
            feedback_miss2_3.tStart = t  # local t and not account for scr refresh
            feedback_miss2_3.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(feedback_miss2_3, 'tStartRefresh')  # time at next scr refresh
            feedback_miss2_3.setAutoDraw(True)
        if feedback_miss2_3.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > feedback_miss2_3.tStartRefresh + time_miss-frameTolerance:
                # keep track of stop time/frame for later
                feedback_miss2_3.tStop = t  # not accounting for scr refresh
                feedback_miss2_3.frameNStop = frameN  # exact frame index
                win.timeOnFlip(feedback_miss2_3, 'tStopRefresh')  # time at next scr refresh
                feedback_miss2_3.setAutoDraw(False)
        
        # *new_angry_4* updates
        if new_angry_4.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            new_angry_4.frameNStart = frameN  # exact frame index
            new_angry_4.tStart = t  # local t and not account for scr refresh
            new_angry_4.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(new_angry_4, 'tStartRefresh')  # time at next scr refresh
            new_angry_4.setAutoDraw(True)
        if new_angry_4.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > new_angry_4.tStartRefresh + time_a-frameTolerance:
                # keep track of stop time/frame for later
                new_angry_4.tStop = t  # not accounting for scr refresh
                new_angry_4.frameNStop = frameN  # exact frame index
                win.timeOnFlip(new_angry_4, 'tStopRefresh')  # time at next scr refresh
                new_angry_4.setAutoDraw(False)
        
        # *new_happy_4* updates
        if new_happy_4.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            new_happy_4.frameNStart = frameN  # exact frame index
            new_happy_4.tStart = t  # local t and not account for scr refresh
            new_happy_4.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(new_happy_4, 'tStartRefresh')  # time at next scr refresh
            new_happy_4.setAutoDraw(True)
        if new_happy_4.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > new_happy_4.tStartRefresh + time_h-frameTolerance:
                # keep track of stop time/frame for later
                new_happy_4.tStop = t  # not accounting for scr refresh
                new_happy_4.frameNStop = frameN  # exact frame index
                win.timeOnFlip(new_happy_4, 'tStopRefresh')  # time at next scr refresh
                new_happy_4.setAutoDraw(False)
        
        # *left_disp_fb_4* updates
        if left_disp_fb_4.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            left_disp_fb_4.frameNStart = frameN  # exact frame index
            left_disp_fb_4.tStart = t  # local t and not account for scr refresh
            left_disp_fb_4.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(left_disp_fb_4, 'tStartRefresh')  # time at next scr refresh
            left_disp_fb_4.setAutoDraw(True)
        if left_disp_fb_4.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > left_disp_fb_4.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                left_disp_fb_4.tStop = t  # not accounting for scr refresh
                left_disp_fb_4.frameNStop = frameN  # exact frame index
                win.timeOnFlip(left_disp_fb_4, 'tStopRefresh')  # time at next scr refresh
                left_disp_fb_4.setAutoDraw(False)
        
        # *right_disp_fb_4* updates
        if right_disp_fb_4.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            right_disp_fb_4.frameNStart = frameN  # exact frame index
            right_disp_fb_4.tStart = t  # local t and not account for scr refresh
            right_disp_fb_4.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(right_disp_fb_4, 'tStartRefresh')  # time at next scr refresh
            right_disp_fb_4.setAutoDraw(True)
        if right_disp_fb_4.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > right_disp_fb_4.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                right_disp_fb_4.tStop = t  # not accounting for scr refresh
                right_disp_fb_4.frameNStop = frameN  # exact frame index
                win.timeOnFlip(right_disp_fb_4, 'tStopRefresh')  # time at next scr refresh
                right_disp_fb_4.setAutoDraw(False)
        
        # *Stimulus_fb_4* updates
        if Stimulus_fb_4.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            Stimulus_fb_4.frameNStart = frameN  # exact frame index
            Stimulus_fb_4.tStart = t  # local t and not account for scr refresh
            Stimulus_fb_4.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(Stimulus_fb_4, 'tStartRefresh')  # time at next scr refresh
            Stimulus_fb_4.setAutoDraw(True)
        if Stimulus_fb_4.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > Stimulus_fb_4.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                Stimulus_fb_4.tStop = t  # not accounting for scr refresh
                Stimulus_fb_4.frameNStop = frameN  # exact frame index
                win.timeOnFlip(Stimulus_fb_4, 'tStopRefresh')  # time at next scr refresh
                Stimulus_fb_4.setAutoDraw(False)
        
        # *Prob_feedback* updates
        if Prob_feedback.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            Prob_feedback.frameNStart = frameN  # exact frame index
            Prob_feedback.tStart = t  # local t and not account for scr refresh
            Prob_feedback.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(Prob_feedback, 'tStartRefresh')  # time at next scr refresh
            Prob_feedback.setAutoDraw(True)
        if Prob_feedback.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > Prob_feedback.tStartRefresh + 6-frameTolerance:
                # keep track of stop time/frame for later
                Prob_feedback.tStop = t  # not accounting for scr refresh
                Prob_feedback.frameNStop = frameN  # exact frame index
                win.timeOnFlip(Prob_feedback, 'tStopRefresh')  # time at next scr refresh
                Prob_feedback.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in feedback_probComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "feedback_prob"-------
    for thisComponent in feedback_probComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    if (response_training.keys == 'left' and CorrCat == left_cat) or (response_training_2.keys == 'right' and CorrDict[stim_thisN] == right_cat):
        thisExp.addData ('CorrectAns', "Yes")
        print("Correct Answer!")
    else:
        thisExp.addData ('CorrectAns', "No")
        print("Wrong Answer!")
    prob_fb.addData('feedback_miss2_3.started', feedback_miss2_3.tStartRefresh)
    prob_fb.addData('feedback_miss2_3.stopped', feedback_miss2_3.tStopRefresh)
    new_angry_4.stop()
    new_happy_4.stop()
    prob_fb.addData('left_disp_fb_4.started', left_disp_fb_4.tStartRefresh)
    prob_fb.addData('left_disp_fb_4.stopped', left_disp_fb_4.tStopRefresh)
    prob_fb.addData('right_disp_fb_4.started', right_disp_fb_4.tStartRefresh)
    prob_fb.addData('right_disp_fb_4.stopped', right_disp_fb_4.tStopRefresh)
    prob_fb.addData('Stimulus_fb_4.started', Stimulus_fb_4.tStartRefresh)
    prob_fb.addData('Stimulus_fb_4.stopped', Stimulus_fb_4.tStopRefresh)
    prob_fb.addData('Prob_feedback.started', Prob_feedback.tStartRefresh)
    prob_fb.addData('Prob_feedback.stopped', Prob_feedback.tStopRefresh)
    # the Routine "feedback_prob" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    thisExp.nextEntry()
    
# completed 1 repeats of 'prob_fb'


# ------Prepare to start Routine "Alles_Klar_3"-------
continueRoutine = True
# update component parameters for each repeat
Allesklar_resp.keys = []
Allesklar_resp.rt = []
_Allesklar_resp_allKeys = []
# keep track of which components have finished
Alles_Klar_3Components = [Allesklartext, Allesklar_resp]
for thisComponent in Alles_Klar_3Components:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
Alles_Klar_3Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "Alles_Klar_3"-------
while continueRoutine:
    # get current time
    t = Alles_Klar_3Clock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=Alles_Klar_3Clock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *Allesklartext* updates
    if Allesklartext.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        Allesklartext.frameNStart = frameN  # exact frame index
        Allesklartext.tStart = t  # local t and not account for scr refresh
        Allesklartext.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(Allesklartext, 'tStartRefresh')  # time at next scr refresh
        Allesklartext.setAutoDraw(True)
    
    # *Allesklar_resp* updates
    waitOnFlip = False
    if Allesklar_resp.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        Allesklar_resp.frameNStart = frameN  # exact frame index
        Allesklar_resp.tStart = t  # local t and not account for scr refresh
        Allesklar_resp.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(Allesklar_resp, 'tStartRefresh')  # time at next scr refresh
        Allesklar_resp.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(Allesklar_resp.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(Allesklar_resp.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if Allesklar_resp.status == STARTED and not waitOnFlip:
        theseKeys = Allesklar_resp.getKeys(keyList=None, waitRelease=False)
        _Allesklar_resp_allKeys.extend(theseKeys)
        if len(_Allesklar_resp_allKeys):
            Allesklar_resp.keys = _Allesklar_resp_allKeys[-1].name  # just the last key pressed
            Allesklar_resp.rt = _Allesklar_resp_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in Alles_Klar_3Components:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "Alles_Klar_3"-------
for thisComponent in Alles_Klar_3Components:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('Allesklartext.started', Allesklartext.tStartRefresh)
thisExp.addData('Allesklartext.stopped', Allesklartext.tStopRefresh)
# the Routine "Alles_Klar_3" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
blocks = data.TrialHandler(nReps=4, method='random', 
    extraInfo=expInfo, originPath=-1,
    trialList=[None],
    seed=None, name='blocks')
thisExp.addLoop(blocks)  # add the loop to the experiment
thisBlock = blocks.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb = thisBlock.rgb)
if thisBlock != None:
    for paramName in thisBlock:
        exec('{} = thisBlock[paramName]'.format(paramName))

for thisBlock in blocks:
    currentLoop = blocks
    # abbreviate parameter names if possible (e.g. rgb = thisBlock.rgb)
    if thisBlock != None:
        for paramName in thisBlock:
            exec('{} = thisBlock[paramName]'.format(paramName))
    
    # ------Prepare to start Routine "BlockCode"-------
    continueRoutine = True
    # update component parameters for each repeat
    #Define filenames according to Block
    if Blocklist[blocks.thisRepN]== "social":
        #trial_file = "trials_social.xlsx"
        filenames_noxl = [
            ["ADFES/Freigestellt/Neutral/F01-Neutral-Face Forward_freigestellt.mp4", "ADFES/Freigestellt/Joy/Used/F01-Joy-Face Forward_freigestellt.mp4", "ADFES/Freigestellt/Anger/F01-Anger-Face Forward_freigestellt.mp4"],
            ["ADFES/Freigestellt/Neutral/F04-Neutral-Face Forward_freigestellt.mp4", "ADFES/Freigestellt/Joy/Used/F04-Joy-Face Forward_freigestellt.mp4", "ADFES/Freigestellt/Anger/F04-Anger-Face Forward_freigestellt.mp4"],
            ["ADFES/Freigestellt/Neutral/M06-Neutral-Face Forward_freigestellt.mp4", "ADFES/Freigestellt/Joy/Used/M06-Joy-Face Forward_freigestellt.mp4", "ADFES/Freigestellt/Anger/M06-Anger-Face Forward_freigestellt.mp4"],
            ["ADFES/Freigestellt/Neutral/M02-Neutral-Face Forward_freigestellt.mp4", "ADFES/Freigestellt/Joy/Used/M02-Joy-Face Forward_freigestellt.mp4", "ADFES/Freigestellt/Anger/M02-Anger-Face Forward_freigestellt.mp4"]
        ]
    elif Blocklist[blocks.thisRepN] == "nonsocial":
        #trial_file = "trials_nonsocial.xlsx"
        filenames_noxl = [
            ["Mandalas_new/neutral/Mandala3_neutral.mp4", "Mandalas_new/happy/Mandala3_happy.mp4", "Mandalas_new/angry/Mandala3_angry.mp4"],
            ["Mandalas_new/neutral/Mandala2_neutral.mp4", "Mandalas_new/happy/Mandala2_happy.mp4", "Mandalas_new/angry/Mandala2_angry.mp4"],
            ["Mandalas_new/neutral/Mandala4_neutral.mp4", "Mandalas_new/happy/Mandala4_happy.mp4", "Mandalas_new/angry/Mandala4_angry.mp4"],
            ["Mandalas_new/neutral/Mandala1_neutral.mp4", "Mandalas_new/happy/Mandala1_happy.mp4", "Mandalas_new/angry/Mandala1_angry.mp4"]
        ]
    #Load Excel file (not relevant anymore)
    '''path = os.path.join(cwd, trial_file)
    book = openpyxl.load_workbook(filename = path)
    sheet = book.active
    
    listAB = ['A', 'B']*2
    random.shuffle(listAB)
    
    for index, element in enumerate(listAB, start=1):
        sheet.cell(row = index + 1, column = 5).value = element
    
    filenames = []
    for x in range(4):
        filenames.append([])
    
    #here you iterate over the rows in the specific column
    for row in range(2,6):
        for column in "BCD":  #Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            filenames[row-2].append(sheet[cell_name].value) # the value of the specific cell
    
    randnumbers = []
    for x in range(1, 5):
        randnumbers.append(random.randint(9, 99))
    
    for row in range(2,6):
        for column in "A":  #Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            sheet[cell_name].value = randnumbers[row-2]# the value of the specific cell
    
    book.save(filename = trial_file)'''
    
    # keep track of which components have finished
    BlockCodeComponents = []
    for thisComponent in BlockCodeComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    BlockCodeClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "BlockCode"-------
    while continueRoutine:
        # get current time
        t = BlockCodeClock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=BlockCodeClock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in BlockCodeComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "BlockCode"-------
    for thisComponent in BlockCodeComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # the Routine "BlockCode" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # set up handler to look after randomisation of conditions etc
    cycles = data.TrialHandler(nReps=6, method='sequential', 
        extraInfo=expInfo, originPath=-1,
        trialList=[None],
        seed=None, name='cycles')
    thisExp.addLoop(cycles)  # add the loop to the experiment
    thisCycle = cycles.trialList[0]  # so we can initialise stimuli with some values
    # abbreviate parameter names if possible (e.g. rgb = thisCycle.rgb)
    if thisCycle != None:
        for paramName in thisCycle:
            exec('{} = thisCycle[paramName]'.format(paramName))
    
    for thisCycle in cycles:
        currentLoop = cycles
        # abbreviate parameter names if possible (e.g. rgb = thisCycle.rgb)
        if thisCycle != None:
            for paramName in thisCycle:
                exec('{} = thisCycle[paramName]'.format(paramName))
        
        # ------Prepare to start Routine "LateralizationByCycle"-------
        continueRoutine = True
        # update component parameters for each repeat
        #randomize lateralization of A and B each cycle
        if choice([True, False]):
            right_cat = "A"
            left_cat = "B"
        else:
            left_cat = "A"
            right_cat = "B"
            
        #shuffe stimuli per cycle
        random.shuffle(numbarray[blocks.thisRepN])
        print("numbarray[blocks.thisRepN]:", numbarray[blocks.thisRepN])
        #shuffle filenames per cycle
        random.shuffle(filenames_noxl)
        print(filenames_noxl)
        
        ###IRRELEVANT
        #randomize picture/number association each cycle
        '''
        random.shuffle(filenames)
        
        for row in range(2,6):
            index = 0
            for column in "BCD":  #Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                sheet[cell_name].value = filenames[row-2][index]# the value of the specific cell
                index += 1
        
        book.save(filename = trial_file)
        '''
        
        #Define Text, dependent on Block and Cycle
        if blocks.thisN == 0 and cycles.thisRepN == 0:
            CycleText = centerfy(
            'blocks:' + str(blocks.thisN) + ', cycles:' + str(cycles.thisRepN) + '\n'
            'right_cat: ' + str(right_cat) + ' left_cat: ' + str(left_cat) + '\n\n'
            'Wir beginnen nun mit dem ersten Block.\n\n' 
            'Alles funktioniert genau wie in den Trainingsdurchläufen.\n\n'
            'Im ersten Durchgang müssen Sie \n die Zuordnung mit den Pfeiltasten raten.\n\n'
            'Versuchen Sie aber, sich die richtige Zuordnung \n '
            'genau einzuprägen!\n'
            'Insgesamt werden sie 4 verschiedene Zahlen sehen,\n danach gibt es eine kurze Pause.\n\n'
            'Wenn Sie keine Fragen mehr haben, \n drücken Sie eine beliebige Taste, \n um den ersten Block zu starten!\n'
            )
        elif blocks.thisN in [0, 1, 2, 3] and cycles.thisRepN == 1:
            CycleText = centerfy(
            'blocks:' + str(blocks.thisN) + ', cycles:' + str(cycles.thisRepN) + '\n'
            'right_cat: ' + str(right_cat) + ' left_cat: ' + str(left_cat) + '\n\n'
            'Sie haben nun alle Zahlen in diesem Block \n ein erstes Mal gesehen.\n\n'
            'Machen Sie kurz Pause!\n\n'
            'Sie werden gleich erneut alle Zahlen gezeigt bekommen.\n' 
            'Versuchen Sie, die Zahlen richtig zuzuordnen.\n\n'
            'Beachten Sie, dass Bilder und Zahlen\n' 
            'nun neu kombiniert wurden.\n\n'
            'Drücken Sie eine beliebige Taste,\n'
            'um den nächsten Durchgang zu starten.'
            )
        elif blocks.thisN in [0, 1, 2, 3] and cycles.thisRepN in [2, 3, 4, 5, 6, 7, 8]:
            CycleText = centerfy(
            'blocks:' + str(blocks.thisN) + ', cycles:' + str(cycles.thisRepN) + '\n'
            'right_cat: ' + str(right_cat) + ' left_cat: ' + str(left_cat) + '\n\n'
            'Sehr gut!\n\n'
            'Sie haben einen weiteren Durchgang abgeschlossen!\n\n'
            'Es folgen noch ' + str(cycles.nRemaining + 1) + ' weitere Durchgänge,\n'
            'in denen Sie wieder jeweils alle Zahlen gezeigt bekommen.\n\n'
            'Drücken Sie eine beliebige Taste, \n'
            'um den nächsten Durchgang zu starten.'
            )
        elif blocks.thisN in [1, 2, 3] and cycles.thisRepN == 0:
            CycleText = centerfy(
            'blocks:' + str(blocks.thisN) + ', cycles:' + str(cycles.thisRepN) + '\n'
            'right_cat: ' + str(right_cat) + ' left_cat: ' + str(left_cat) + '\n\n'
            'Sie haben nun einen  vollständigen Block abgeschlossen.\n\n'
            'Im nächsten Block werden andere Bilder angezeigt.\n'
            'Statt Fraktalen werden Sie nun Gesichter sehen\n'
            '(oder umgekehrt).\n\n'
            'Das Prinzip bleibt gleich.\n\n'
            'Es folgen noch ' + str(blocks.nRemaining + 1) + ' weitere Blöcke.\n\n'
            'Wenn Sie bereit sind, \n drücken Sie eine beliebige Taste,\n'
            'um den nächsten Block zu starten.\n'
            )
        CycleText3.setText('PAUSE\n\nPress any key, to start the next cycle.')
        key_resp_6.keys = []
        key_resp_6.rt = []
        _key_resp_6_allKeys = []
        # keep track of which components have finished
        LateralizationByCycleComponents = [CycleText1, CycleText2, CycleText3, CycleText4, key_resp_6]
        for thisComponent in LateralizationByCycleComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        LateralizationByCycleClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
        frameN = -1
        
        # -------Run Routine "LateralizationByCycle"-------
        while continueRoutine:
            # get current time
            t = LateralizationByCycleClock.getTime()
            tThisFlip = win.getFutureFlipTime(clock=LateralizationByCycleClock)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *CycleText1* updates
            if CycleText1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                CycleText1.frameNStart = frameN  # exact frame index
                CycleText1.tStart = t  # local t and not account for scr refresh
                CycleText1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(CycleText1, 'tStartRefresh')  # time at next scr refresh
                CycleText1.setAutoDraw(True)
            if CycleText1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > CycleText1.tStartRefresh + CycleText1_dur-frameTolerance:
                    # keep track of stop time/frame for later
                    CycleText1.tStop = t  # not accounting for scr refresh
                    CycleText1.frameNStop = frameN  # exact frame index
                    win.timeOnFlip(CycleText1, 'tStopRefresh')  # time at next scr refresh
                    CycleText1.setAutoDraw(False)
            
            # *CycleText2* updates
            if CycleText2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                CycleText2.frameNStart = frameN  # exact frame index
                CycleText2.tStart = t  # local t and not account for scr refresh
                CycleText2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(CycleText2, 'tStartRefresh')  # time at next scr refresh
                CycleText2.setAutoDraw(True)
            if CycleText2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > CycleText2.tStartRefresh + CycleText2_dur-frameTolerance:
                    # keep track of stop time/frame for later
                    CycleText2.tStop = t  # not accounting for scr refresh
                    CycleText2.frameNStop = frameN  # exact frame index
                    win.timeOnFlip(CycleText2, 'tStopRefresh')  # time at next scr refresh
                    CycleText2.setAutoDraw(False)
            
            # *CycleText3* updates
            if CycleText3.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                CycleText3.frameNStart = frameN  # exact frame index
                CycleText3.tStart = t  # local t and not account for scr refresh
                CycleText3.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(CycleText3, 'tStartRefresh')  # time at next scr refresh
                CycleText3.setAutoDraw(True)
            if CycleText3.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > CycleText3.tStartRefresh + CycleText3_dur-frameTolerance:
                    # keep track of stop time/frame for later
                    CycleText3.tStop = t  # not accounting for scr refresh
                    CycleText3.frameNStop = frameN  # exact frame index
                    win.timeOnFlip(CycleText3, 'tStopRefresh')  # time at next scr refresh
                    CycleText3.setAutoDraw(False)
            
            # *CycleText4* updates
            if CycleText4.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                CycleText4.frameNStart = frameN  # exact frame index
                CycleText4.tStart = t  # local t and not account for scr refresh
                CycleText4.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(CycleText4, 'tStartRefresh')  # time at next scr refresh
                CycleText4.setAutoDraw(True)
            if CycleText4.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > CycleText4.tStartRefresh + CycleText4_dur-frameTolerance:
                    # keep track of stop time/frame for later
                    CycleText4.tStop = t  # not accounting for scr refresh
                    CycleText4.frameNStop = frameN  # exact frame index
                    win.timeOnFlip(CycleText4, 'tStopRefresh')  # time at next scr refresh
                    CycleText4.setAutoDraw(False)
            
            # *key_resp_6* updates
            waitOnFlip = False
            if key_resp_6.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                key_resp_6.frameNStart = frameN  # exact frame index
                key_resp_6.tStart = t  # local t and not account for scr refresh
                key_resp_6.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(key_resp_6, 'tStartRefresh')  # time at next scr refresh
                key_resp_6.status = STARTED
                # keyboard checking is just starting
                waitOnFlip = True
                win.callOnFlip(key_resp_6.clock.reset)  # t=0 on next screen flip
                win.callOnFlip(key_resp_6.clearEvents, eventType='keyboard')  # clear events on next screen flip
            if key_resp_6.status == STARTED and not waitOnFlip:
                theseKeys = key_resp_6.getKeys(keyList=None, waitRelease=False)
                _key_resp_6_allKeys.extend(theseKeys)
                if len(_key_resp_6_allKeys):
                    key_resp_6.keys = _key_resp_6_allKeys[-1].name  # just the last key pressed
                    key_resp_6.rt = _key_resp_6_allKeys[-1].rt
                    # a response ends the routine
                    continueRoutine = False
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in LateralizationByCycleComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # -------Ending Routine "LateralizationByCycle"-------
        for thisComponent in LateralizationByCycleComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        cycles.addData('CycleText1.started', CycleText1.tStartRefresh)
        cycles.addData('CycleText1.stopped', CycleText1.tStopRefresh)
        cycles.addData('CycleText2.started', CycleText2.tStartRefresh)
        cycles.addData('CycleText2.stopped', CycleText2.tStopRefresh)
        cycles.addData('CycleText3.started', CycleText3.tStartRefresh)
        cycles.addData('CycleText3.stopped', CycleText3.tStopRefresh)
        cycles.addData('CycleText4.started', CycleText4.tStartRefresh)
        cycles.addData('CycleText4.stopped', CycleText4.tStopRefresh)
        # the Routine "LateralizationByCycle" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        # set up handler to look after randomisation of conditions etc
        trials = data.TrialHandler(nReps=4, method='sequential', 
            extraInfo=expInfo, originPath=-1,
            trialList=[None],
            seed=None, name='trials')
        thisExp.addLoop(trials)  # add the loop to the experiment
        thisTrial = trials.trialList[0]  # so we can initialise stimuli with some values
        # abbreviate parameter names if possible (e.g. rgb = thisTrial.rgb)
        if thisTrial != None:
            for paramName in thisTrial:
                exec('{} = thisTrial[paramName]'.format(paramName))
        
        for thisTrial in trials:
            currentLoop = trials
            # abbreviate parameter names if possible (e.g. rgb = thisTrial.rgb)
            if thisTrial != None:
                for paramName in thisTrial:
                    exec('{} = thisTrial[paramName]'.format(paramName))
            
            # ------Prepare to start Routine "trial_2"-------
            continueRoutine = True
            routineTimer.add(7.000000)
            # update component parameters for each repeat
            filename_thisN_neutral = filenames_noxl[trials.thisRepN][0]
            filename_thisN_happy = filenames_noxl[trials.thisRepN][1]
            filename_thisN_angry = filenames_noxl[trials.thisRepN][2]
            stim_thisN = numbarray[blocks.thisRepN][trials.thisRepN]
            print(stim_thisN)
            print(CorrDict[stim_thisN])
            new_neutral_2_jpg.setImage(filename_thisN_neutral)
            response_training_2.keys = []
            response_training_2.rt = []
            _response_training_2_allKeys = []
            Stimulus_2.setText(str(int(stim_thisN)))
            left_disp_2.setText(left_cat)
            right_disp_2.setText(right_cat)
            # keep track of which components have finished
            trial_2Components = [fix_cross_2, new_neutral_2_jpg, response_training_2, Stimulus_2, left_disp_2, right_disp_2]
            for thisComponent in trial_2Components:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            trial_2Clock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
            frameN = -1
            
            # -------Run Routine "trial_2"-------
            while continueRoutine and routineTimer.getTime() > 0:
                # get current time
                t = trial_2Clock.getTime()
                tThisFlip = win.getFutureFlipTime(clock=trial_2Clock)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *fix_cross_2* updates
                if fix_cross_2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                    # keep track of start time/frame for later
                    fix_cross_2.frameNStart = frameN  # exact frame index
                    fix_cross_2.tStart = t  # local t and not account for scr refresh
                    fix_cross_2.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(fix_cross_2, 'tStartRefresh')  # time at next scr refresh
                    fix_cross_2.setAutoDraw(True)
                if fix_cross_2.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > fix_cross_2.tStartRefresh + 1-frameTolerance:
                        # keep track of stop time/frame for later
                        fix_cross_2.tStop = t  # not accounting for scr refresh
                        fix_cross_2.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(fix_cross_2, 'tStopRefresh')  # time at next scr refresh
                        fix_cross_2.setAutoDraw(False)
                
                # *new_neutral_2_jpg* updates
                if new_neutral_2_jpg.status == NOT_STARTED and tThisFlip >= 1.0-frameTolerance:
                    # keep track of start time/frame for later
                    new_neutral_2_jpg.frameNStart = frameN  # exact frame index
                    new_neutral_2_jpg.tStart = t  # local t and not account for scr refresh
                    new_neutral_2_jpg.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(new_neutral_2_jpg, 'tStartRefresh')  # time at next scr refresh
                    new_neutral_2_jpg.setAutoDraw(True)
                if new_neutral_2_jpg.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > new_neutral_2_jpg.tStartRefresh + 6.0-frameTolerance:
                        # keep track of stop time/frame for later
                        new_neutral_2_jpg.tStop = t  # not accounting for scr refresh
                        new_neutral_2_jpg.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(new_neutral_2_jpg, 'tStopRefresh')  # time at next scr refresh
                        new_neutral_2_jpg.setAutoDraw(False)
                
                # *response_training_2* updates
                waitOnFlip = False
                if response_training_2.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
                    # keep track of start time/frame for later
                    response_training_2.frameNStart = frameN  # exact frame index
                    response_training_2.tStart = t  # local t and not account for scr refresh
                    response_training_2.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(response_training_2, 'tStartRefresh')  # time at next scr refresh
                    response_training_2.status = STARTED
                    # keyboard checking is just starting
                    waitOnFlip = True
                    win.callOnFlip(response_training_2.clock.reset)  # t=0 on next screen flip
                    win.callOnFlip(response_training_2.clearEvents, eventType='keyboard')  # clear events on next screen flip
                if response_training_2.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > response_training_2.tStartRefresh + 6-frameTolerance:
                        # keep track of stop time/frame for later
                        response_training_2.tStop = t  # not accounting for scr refresh
                        response_training_2.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(response_training_2, 'tStopRefresh')  # time at next scr refresh
                        response_training_2.status = FINISHED
                if response_training_2.status == STARTED and not waitOnFlip:
                    theseKeys = response_training_2.getKeys(keyList=['left', 'right'], waitRelease=False)
                    _response_training_2_allKeys.extend(theseKeys)
                    if len(_response_training_2_allKeys):
                        response_training_2.keys = _response_training_2_allKeys[-1].name  # just the last key pressed
                        response_training_2.rt = _response_training_2_allKeys[-1].rt
                        # was this correct?
                        if (response_training_2.keys == str(CorrCat)) or (response_training_2.keys == CorrCat):
                            response_training_2.corr = 1
                        else:
                            response_training_2.corr = 0
                        # a response ends the routine
                        continueRoutine = False
                
                # *Stimulus_2* updates
                if Stimulus_2.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
                    # keep track of start time/frame for later
                    Stimulus_2.frameNStart = frameN  # exact frame index
                    Stimulus_2.tStart = t  # local t and not account for scr refresh
                    Stimulus_2.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(Stimulus_2, 'tStartRefresh')  # time at next scr refresh
                    Stimulus_2.setAutoDraw(True)
                if Stimulus_2.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > Stimulus_2.tStartRefresh + 6-frameTolerance:
                        # keep track of stop time/frame for later
                        Stimulus_2.tStop = t  # not accounting for scr refresh
                        Stimulus_2.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(Stimulus_2, 'tStopRefresh')  # time at next scr refresh
                        Stimulus_2.setAutoDraw(False)
                
                # *left_disp_2* updates
                if left_disp_2.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
                    # keep track of start time/frame for later
                    left_disp_2.frameNStart = frameN  # exact frame index
                    left_disp_2.tStart = t  # local t and not account for scr refresh
                    left_disp_2.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(left_disp_2, 'tStartRefresh')  # time at next scr refresh
                    left_disp_2.setAutoDraw(True)
                if left_disp_2.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > left_disp_2.tStartRefresh + 6-frameTolerance:
                        # keep track of stop time/frame for later
                        left_disp_2.tStop = t  # not accounting for scr refresh
                        left_disp_2.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(left_disp_2, 'tStopRefresh')  # time at next scr refresh
                        left_disp_2.setAutoDraw(False)
                
                # *right_disp_2* updates
                if right_disp_2.status == NOT_STARTED and tThisFlip >= 1-frameTolerance:
                    # keep track of start time/frame for later
                    right_disp_2.frameNStart = frameN  # exact frame index
                    right_disp_2.tStart = t  # local t and not account for scr refresh
                    right_disp_2.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(right_disp_2, 'tStartRefresh')  # time at next scr refresh
                    right_disp_2.setAutoDraw(True)
                if right_disp_2.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > right_disp_2.tStartRefresh + 6-frameTolerance:
                        # keep track of stop time/frame for later
                        right_disp_2.tStop = t  # not accounting for scr refresh
                        right_disp_2.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(right_disp_2, 'tStopRefresh')  # time at next scr refresh
                        right_disp_2.setAutoDraw(False)
                
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in trial_2Components:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # -------Ending Routine "trial_2"-------
            for thisComponent in trial_2Components:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            #Speichern von keyangry und keyhappy ins Data file in jedem Trial
            thisExp.addData ('left', left_cat)
            thisExp.addData ('right', right_cat)
            thisExp.addData('stim', stim_thisN)
            
            if response_training_2.keys == 'left':
                thisExp.addData('ParticipantAnswer', left_cat)
                print("Participant responded with:", left_cat)
            elif response_training_2.keys == 'right':
                thisExp.addData('ParticipantAnswer', right_cat)
                print("Participant responded with:", right_cat)
            elif response_training_2.keys == None:
                thisExp.addData('ParticipantAnswer', None)
                print("Participant responded with:", None)
            trials.addData('fix_cross_2.started', fix_cross_2.tStartRefresh)
            trials.addData('fix_cross_2.stopped', fix_cross_2.tStopRefresh)
            trials.addData('new_neutral_2_jpg.started', new_neutral_2_jpg.tStartRefresh)
            trials.addData('new_neutral_2_jpg.stopped', new_neutral_2_jpg.tStopRefresh)
            # check responses
            if response_training_2.keys in ['', [], None]:  # No response was made
                response_training_2.keys = None
                # was no response the correct answer?!
                if str(CorrCat).lower() == 'none':
                   response_training_2.corr = 1;  # correct non-response
                else:
                   response_training_2.corr = 0;  # failed to respond (incorrectly)
            # store data for trials (TrialHandler)
            trials.addData('response_training_2.keys',response_training_2.keys)
            trials.addData('response_training_2.corr', response_training_2.corr)
            if response_training_2.keys != None:  # we had a response
                trials.addData('response_training_2.rt', response_training_2.rt)
            trials.addData('response_training_2.started', response_training_2.tStartRefresh)
            trials.addData('response_training_2.stopped', response_training_2.tStopRefresh)
            trials.addData('Stimulus_2.started', Stimulus_2.tStartRefresh)
            trials.addData('Stimulus_2.stopped', Stimulus_2.tStopRefresh)
            trials.addData('left_disp_2.started', left_disp_2.tStartRefresh)
            trials.addData('left_disp_2.stopped', left_disp_2.tStopRefresh)
            trials.addData('right_disp_2.started', right_disp_2.tStartRefresh)
            trials.addData('right_disp_2.stopped', right_disp_2.tStopRefresh)
            
            # ------Prepare to start Routine "feedback"-------
            continueRoutine = True
            # update component parameters for each repeat
            #Feedback Zuweisung
            # to make feedback probabilistic: uncomment nested if clause
            # this will show happy for correct (congruent) in 80% of cases 
            # and angry for correct (incongruent) in 20% of cases
            # probability can be changed by manipulating the sequence in p
            
            
            #if correct show happy video (with 0.8 probability)
            if (response_training_2.keys == 'left' and CorrDict[stim_thisN] == left_cat) or (response_training_2.keys == 'right' and CorrDict[stim_thisN] == right_cat):
                if choice([True, False], p = [0.8, 0.2]):
                    time_h = 6.0
                    time_a = 0.0
                else:
                    time_h = 0.0
                    time_a = 6.0
            
            #if not correct show angry video (with 0.8 probability)
            else:
                if choice([True, False], p = [0.8, 0.2]):
                    time_h = 0.0
                    time_a = 6.0
                else:
                    time_h = 6.0
                    time_a = 0.0
            
            #Set color of A/B on screen after response
            time_miss = 0
            if response_training_2.keys == 'left':
                left_color = 'black'
                right_color = 'grey'
            elif response_training_2.keys == 'right':
                left_color = 'grey'
                right_color = 'black'
            elif response_training_2.keys == None:
                left_color = 'grey'
                right_color = 'grey'
                time_miss = 6
                time_a = 0
                time_h = 0
            new_angry = visual.MovieStim3(
                win=win, name='new_angry',units='height', 
                noAudio = True,
                filename=filename_thisN_angry,
                ori=0, pos=[0, 0.225], opacity=1,
                loop=False,
                size=[0.5, 0.5],
                depth=-2.0,
                )
            new_happy = visual.MovieStim3(
                win=win, name='new_happy',units='height', 
                noAudio = False,
                filename=filename_thisN_happy,
                ori=0, pos=[0, 0.225], opacity=1,
                loop=False,
                size=[0.5, 0.5],
                depth=-3.0,
                )
            left_disp_fb.setColor(left_color, colorSpace='rgb')
            left_disp_fb.setText(left_cat)
            right_disp_fb.setColor(right_color, colorSpace='rgb')
            right_disp_fb.setText(right_cat)
            Stimulus_fb.setText(str(int(stim_thisN)))
            # keep track of which components have finished
            feedbackComponents = [feedback_miss2, new_angry, new_happy, left_disp_fb, right_disp_fb, Stimulus_fb]
            for thisComponent in feedbackComponents:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            feedbackClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
            frameN = -1
            
            # -------Run Routine "feedback"-------
            while continueRoutine:
                # get current time
                t = feedbackClock.getTime()
                tThisFlip = win.getFutureFlipTime(clock=feedbackClock)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *feedback_miss2* updates
                if feedback_miss2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    feedback_miss2.frameNStart = frameN  # exact frame index
                    feedback_miss2.tStart = t  # local t and not account for scr refresh
                    feedback_miss2.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(feedback_miss2, 'tStartRefresh')  # time at next scr refresh
                    feedback_miss2.setAutoDraw(True)
                if feedback_miss2.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > feedback_miss2.tStartRefresh + time_miss-frameTolerance:
                        # keep track of stop time/frame for later
                        feedback_miss2.tStop = t  # not accounting for scr refresh
                        feedback_miss2.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(feedback_miss2, 'tStopRefresh')  # time at next scr refresh
                        feedback_miss2.setAutoDraw(False)
                
                # *new_angry* updates
                if new_angry.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    new_angry.frameNStart = frameN  # exact frame index
                    new_angry.tStart = t  # local t and not account for scr refresh
                    new_angry.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(new_angry, 'tStartRefresh')  # time at next scr refresh
                    new_angry.setAutoDraw(True)
                if new_angry.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > new_angry.tStartRefresh + time_a-frameTolerance:
                        # keep track of stop time/frame for later
                        new_angry.tStop = t  # not accounting for scr refresh
                        new_angry.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(new_angry, 'tStopRefresh')  # time at next scr refresh
                        new_angry.setAutoDraw(False)
                
                # *new_happy* updates
                if new_happy.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    new_happy.frameNStart = frameN  # exact frame index
                    new_happy.tStart = t  # local t and not account for scr refresh
                    new_happy.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(new_happy, 'tStartRefresh')  # time at next scr refresh
                    new_happy.setAutoDraw(True)
                if new_happy.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > new_happy.tStartRefresh + time_h-frameTolerance:
                        # keep track of stop time/frame for later
                        new_happy.tStop = t  # not accounting for scr refresh
                        new_happy.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(new_happy, 'tStopRefresh')  # time at next scr refresh
                        new_happy.setAutoDraw(False)
                
                # *left_disp_fb* updates
                if left_disp_fb.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                    # keep track of start time/frame for later
                    left_disp_fb.frameNStart = frameN  # exact frame index
                    left_disp_fb.tStart = t  # local t and not account for scr refresh
                    left_disp_fb.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(left_disp_fb, 'tStartRefresh')  # time at next scr refresh
                    left_disp_fb.setAutoDraw(True)
                if left_disp_fb.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > left_disp_fb.tStartRefresh + 6-frameTolerance:
                        # keep track of stop time/frame for later
                        left_disp_fb.tStop = t  # not accounting for scr refresh
                        left_disp_fb.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(left_disp_fb, 'tStopRefresh')  # time at next scr refresh
                        left_disp_fb.setAutoDraw(False)
                
                # *right_disp_fb* updates
                if right_disp_fb.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                    # keep track of start time/frame for later
                    right_disp_fb.frameNStart = frameN  # exact frame index
                    right_disp_fb.tStart = t  # local t and not account for scr refresh
                    right_disp_fb.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(right_disp_fb, 'tStartRefresh')  # time at next scr refresh
                    right_disp_fb.setAutoDraw(True)
                if right_disp_fb.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > right_disp_fb.tStartRefresh + 6-frameTolerance:
                        # keep track of stop time/frame for later
                        right_disp_fb.tStop = t  # not accounting for scr refresh
                        right_disp_fb.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(right_disp_fb, 'tStopRefresh')  # time at next scr refresh
                        right_disp_fb.setAutoDraw(False)
                
                # *Stimulus_fb* updates
                if Stimulus_fb.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                    # keep track of start time/frame for later
                    Stimulus_fb.frameNStart = frameN  # exact frame index
                    Stimulus_fb.tStart = t  # local t and not account for scr refresh
                    Stimulus_fb.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(Stimulus_fb, 'tStartRefresh')  # time at next scr refresh
                    Stimulus_fb.setAutoDraw(True)
                if Stimulus_fb.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > Stimulus_fb.tStartRefresh + 6-frameTolerance:
                        # keep track of stop time/frame for later
                        Stimulus_fb.tStop = t  # not accounting for scr refresh
                        Stimulus_fb.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(Stimulus_fb, 'tStopRefresh')  # time at next scr refresh
                        Stimulus_fb.setAutoDraw(False)
                
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in feedbackComponents:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # -------Ending Routine "feedback"-------
            for thisComponent in feedbackComponents:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            if (response_training_2.keys == 'left' and CorrDict[stim_thisN] == left_cat) or (response_training_2.keys == 'right' and CorrDict[stim_thisN] == right_cat):
                thisExp.addData ('CorrectAns', "Yes")
                print("Correct Answer!")
            else:
                thisExp.addData ('CorrectAns', "No")
                print("Wrong Answer!")
            trials.addData('feedback_miss2.started', feedback_miss2.tStartRefresh)
            trials.addData('feedback_miss2.stopped', feedback_miss2.tStopRefresh)
            new_angry.stop()
            new_happy.stop()
            trials.addData('left_disp_fb.started', left_disp_fb.tStartRefresh)
            trials.addData('left_disp_fb.stopped', left_disp_fb.tStopRefresh)
            trials.addData('right_disp_fb.started', right_disp_fb.tStartRefresh)
            trials.addData('right_disp_fb.stopped', right_disp_fb.tStopRefresh)
            trials.addData('Stimulus_fb.started', Stimulus_fb.tStartRefresh)
            trials.addData('Stimulus_fb.stopped', Stimulus_fb.tStopRefresh)
            # the Routine "feedback" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            
            # ------Prepare to start Routine "Intertrial_Interval"-------
            continueRoutine = True
            routineTimer.add(1.000000)
            # update component parameters for each repeat
            print(trials.thisN)
            # keep track of which components have finished
            Intertrial_IntervalComponents = [text_2]
            for thisComponent in Intertrial_IntervalComponents:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            Intertrial_IntervalClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
            frameN = -1
            
            # -------Run Routine "Intertrial_Interval"-------
            while continueRoutine and routineTimer.getTime() > 0:
                # get current time
                t = Intertrial_IntervalClock.getTime()
                tThisFlip = win.getFutureFlipTime(clock=Intertrial_IntervalClock)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *text_2* updates
                if text_2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    text_2.frameNStart = frameN  # exact frame index
                    text_2.tStart = t  # local t and not account for scr refresh
                    text_2.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(text_2, 'tStartRefresh')  # time at next scr refresh
                    text_2.setAutoDraw(True)
                if text_2.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > text_2.tStartRefresh + 1.0-frameTolerance:
                        # keep track of stop time/frame for later
                        text_2.tStop = t  # not accounting for scr refresh
                        text_2.frameNStop = frameN  # exact frame index
                        win.timeOnFlip(text_2, 'tStopRefresh')  # time at next scr refresh
                        text_2.setAutoDraw(False)
                
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in Intertrial_IntervalComponents:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # -------Ending Routine "Intertrial_Interval"-------
            for thisComponent in Intertrial_IntervalComponents:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            trials.addData('text_2.started', text_2.tStartRefresh)
            trials.addData('text_2.stopped', text_2.tStopRefresh)
            thisExp.nextEntry()
            
        # completed 4 repeats of 'trials'
        
        
        # ------Prepare to start Routine "LatCounter"-------
        continueRoutine = True
        # update component parameters for each repeat
        # keep track of which components have finished
        LatCounterComponents = []
        for thisComponent in LatCounterComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        LatCounterClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
        frameN = -1
        
        # -------Run Routine "LatCounter"-------
        while continueRoutine:
            # get current time
            t = LatCounterClock.getTime()
            tThisFlip = win.getFutureFlipTime(clock=LatCounterClock)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in LatCounterComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # -------Ending Routine "LatCounter"-------
        for thisComponent in LatCounterComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # the Routine "LatCounter" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
    # completed 6 repeats of 'cycles'
    
    
    # ------Prepare to start Routine "BlockCounter"-------
    continueRoutine = True
    # update component parameters for each repeat
    # keep track of which components have finished
    BlockCounterComponents = []
    for thisComponent in BlockCounterComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    BlockCounterClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
    frameN = -1
    
    # -------Run Routine "BlockCounter"-------
    while continueRoutine:
        # get current time
        t = BlockCounterClock.getTime()
        tThisFlip = win.getFutureFlipTime(clock=BlockCounterClock)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in BlockCounterComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # -------Ending Routine "BlockCounter"-------
    for thisComponent in BlockCounterComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # the Routine "BlockCounter" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
# completed 4 repeats of 'blocks'


# ------Prepare to start Routine "Thanks"-------
continueRoutine = True
# update component parameters for each repeat
end.keys = []
end.rt = []
_end_allKeys = []
# keep track of which components have finished
ThanksComponents = [Thank, end]
for thisComponent in ThanksComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
ThanksClock.reset(-_timeToFirstFrame)  # t0 is time of first possible flip
frameN = -1

# -------Run Routine "Thanks"-------
while continueRoutine:
    # get current time
    t = ThanksClock.getTime()
    tThisFlip = win.getFutureFlipTime(clock=ThanksClock)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *Thank* updates
    if Thank.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        Thank.frameNStart = frameN  # exact frame index
        Thank.tStart = t  # local t and not account for scr refresh
        Thank.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(Thank, 'tStartRefresh')  # time at next scr refresh
        Thank.setAutoDraw(True)
    
    # *end* updates
    waitOnFlip = False
    if end.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        end.frameNStart = frameN  # exact frame index
        end.tStart = t  # local t and not account for scr refresh
        end.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(end, 'tStartRefresh')  # time at next scr refresh
        end.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(end.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(end.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if end.status == STARTED and not waitOnFlip:
        theseKeys = end.getKeys(keyList=None, waitRelease=False)
        _end_allKeys.extend(theseKeys)
        if len(_end_allKeys):
            end.keys = _end_allKeys[-1].name  # just the last key pressed
            end.rt = _end_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in ThanksComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# -------Ending Routine "Thanks"-------
for thisComponent in ThanksComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
thisExp.addData('Thank.started', Thank.tStartRefresh)
thisExp.addData('Thank.stopped', Thank.tStopRefresh)
# the Routine "Thanks" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# Flip one final time so any remaining win.callOnFlip() 
# and win.timeOnFlip() tasks get executed before quitting
win.flip()

# these shouldn't be strictly necessary (should auto-save)
thisExp.saveAsWideText(filename+'.csv', delim='auto')
thisExp.saveAsPickle(filename)
# make sure everything is closed down
thisExp.abort()  # or data files will save again on exit
win.close()
core.quit()
